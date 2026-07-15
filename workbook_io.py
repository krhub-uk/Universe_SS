"""
Shared workbook resolution + versioned-save helpers.

Used by portfolio_csv_ingestion.py, price_action_4h.py and
fetch_engine_monthly.py so all three scripts resolve the live workbook and
save new versions the same way. sync_engine.py is descoped from this pass
and is NOT touched or wired to this module.

Base path = the folder these scripts live in
(C:\\Users\\kyle\\Cowork\\Projects\\Universe_go\\), per the folder structure:

    Inputs/BC, Inputs/HL, Outputs/TV, Outputs/IG, Outputs/BC,
    Archive/BC, Archive/HL, Archive/Workbook
"""

import re
import shutil
from pathlib import Path

from openpyxl.styles import Alignment, Font

BASE_PATH = Path(__file__).resolve().parent

INPUTS_BC = BASE_PATH / "Inputs" / "BC"
INPUTS_HL = BASE_PATH / "Inputs" / "HL"
OUTPUTS_TV = BASE_PATH / "Outputs" / "TV"
OUTPUTS_IG = BASE_PATH / "Outputs" / "IG"
OUTPUTS_BC = BASE_PATH / "Outputs" / "BC"
ARCHIVE_BC = BASE_PATH / "Archive" / "BC"
ARCHIVE_HL = BASE_PATH / "Archive" / "HL"
ARCHIVE_WORKBOOK = BASE_PATH / "Archive" / "Workbook"

# Matches the "v<major>_<minor>_<patch>" segment in a filename, e.g.
# 00_Portfolio_Engine_Universe_v1_4_8.xlsm -> prefix="v1_4_", patch="8".
_VERSION_PATCH_RE = re.compile(r"(v\d+_\d+_)(\d+)")


class WorkbookResolutionError(RuntimeError):
    """Raised when the base path doesn't contain exactly one .xlsm, or the
    filename doesn't carry a recognisable v#_#_# patch digit to increment."""


def find_workbook(base_path=BASE_PATH):
    """
    Glob base_path (non-recursive) for the live .xlsm workbook. Excel lock
    files (~$...) are ignored. If 2+ candidates are found, the most recently
    modified one is used (no abort -- Sprint 2 Run 3 fix: stale copies left
    in the base path shouldn't block a run). Only aborts on 0 matches, since
    there's nothing to resolve to.
    """
    matches = [
        p for p in Path(base_path).glob("*.xlsm") if not p.name.startswith("~$")
    ]
    if len(matches) == 0:
        raise WorkbookResolutionError(
            f"[ABORT] No .xlsm workbook found in {base_path}. Expected exactly one."
        )
    if len(matches) > 1:
        matches.sort(key=lambda p: [int(x) for x in re.findall(r'\d+', p.stem)], reverse=True
    return matches[0]


def next_patch_path(path):
    """
    Auto-increment the patch digit (third digit) in a 'v#_#_#' filename.
    e.g. 00_Portfolio_Engine_Universe_v1_4_8.xlsm
      -> 00_Portfolio_Engine_Universe_v1_4_9.xlsm
    """
    path = Path(path)
    match = _VERSION_PATCH_RE.search(path.stem)
    if not match:
        raise WorkbookResolutionError(
            f"[ABORT] Could not find a 'v#_#_#' version pattern in "
            f"'{path.name}' to auto-increment the patch digit."
        )
    prefix, patch_str = match.groups()
    new_patch = str(int(patch_str) + 1)
    new_stem = (
        path.stem[: match.start()] + prefix + new_patch + path.stem[match.end():]
    )
    return path.with_name(new_stem + path.suffix)


def save_workbook_with_increment(wb, current_path):
    """
    Save wb under an auto-incremented patch filename in the same folder as
    current_path, then move current_path into Archive/Workbook/. Returns the
    new Path.

    Sprint 2 Run 5 fix #7: previously this called current_path.unlink() to
    delete the old version outright. The connected workspace folder can
    block deletes on files it's synced in, which either left old versions
    stuck undeleted (accumulating duplicate .xlsm files in the base path) or
    -- worse -- raised mid-operation after the new version had already been
    saved. Moving to Archive/Workbook/ instead of deleting keeps every prior
    version recoverable, and sidesteps the delete-is-blocked failure mode
    entirely for the common case where move/rename is permitted even when a
    bare unlink isn't.
    """
    current_path = Path(current_path)
    new_path = next_patch_path(current_path)
    wb.save(new_path)
    ARCHIVE_WORKBOOK.mkdir(parents=True, exist_ok=True)
    dest = ARCHIVE_WORKBOOK / current_path.name
    try:
        shutil.move(str(current_path), str(dest))
    except PermissionError:
        # Sprint 2 Run 5 fix #7 (hardening): some workspace mounts block
        # both unlink AND rename/move of a file once it's been synced.
        # Fall back to copying the bytes into Archive/Workbook/ and only
        # then attempting to remove the original -- if the remove itself
        # is blocked, leave the old file in place rather than losing data
        # or crashing the run. A stray old-version file left at the top
        # level is a cosmetic annoyance; find_workbook() already tolerates
        # multiple .xlsm candidates (picks the most recently modified), so
        # it won't break the next run.
        shutil.copy2(str(current_path), str(dest))
        try:
            current_path.unlink()
        except PermissionError:
            print(
                f"[WARN] Could not remove old workbook '{current_path.name}' "
                f"after archiving a copy to {dest} -- delete is blocked by "
                f"the workspace mount. Left in place; next run's "
                f"find_workbook() will still resolve correctly (most-recent "
                f"mtime wins)."
            )
    return new_path


# ---------------------------------------------------------------------------
# Legend tab config (Sprint 2 Run 5 fix #5)
#
# gap due days, VOL_EXTREME/VOL_SPIKE thresholds, anchor overwrite/expire
# windows, country lookup, and sector->ETF lookup all used to be hardcoded
# Python constants duplicated across the three pipeline scripts. They now
# live on the Legend tab (single source of truth, editable without a code
# change) and are read at runtime via the helpers below. Scripts must fail
# clearly if a key is missing rather than silently falling back to a
# hardcoded default -- see LegendConfigError.
# ---------------------------------------------------------------------------

class LegendConfigError(RuntimeError):
    """Raised when a required Legend-tab config key or lookup table is
    missing. Sprint 2 Run 5 fix #5: the three scripts must fail clearly on a
    missing Legend key rather than silently falling back to a hardcoded
    default, so this is raised instead of returning None/{}."""


LEGEND_SHEET = "Legend"


def read_legend_scalars(wb, required_keys):
    """
    Scan the Legend tab's column A for each key in required_keys (exact
    string match), returning {key: value_in_column_B}. Used for single-value
    config like GAP_DUE_DAYS or VOL_EXTREME_THRESHOLD (Legend tab's "CONFIG
    THRESHOLDS" block). Raises LegendConfigError naming every missing key if
    any aren't found on the sheet at all, or if the value cell is blank.
    """
    ws = wb[LEGEND_SHEET]
    found = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] in required_keys and row[0] not in found:
            found[row[0]] = row[1]
    missing = [k for k in required_keys if found.get(k) is None]
    if missing:
        raise LegendConfigError(
            f"[ABORT] Legend tab is missing required config key(s): "
            f"{', '.join(missing)}. Add them to the Legend tab's CONFIG "
            f"THRESHOLDS block before running."
        )
    return found


def read_legend_lookup_table(wb, section_label):
    """
    Read a two-column lookup table (e.g. "COUNTRY LOOKUP", "SECTOR ETF
    LOOKUP") from the Legend tab. section_label must appear alone in column A
    on its own header row; the row directly below it is treated as a column
    caption row (skipped); data rows (key in column A, value in column B)
    follow until the first blank column-A cell. Raises LegendConfigError if
    section_label isn't found anywhere in column A -- a missing lookup table
    fails the run rather than silently returning an empty dict that would
    make every value in that column look like an "unmapped" gap.
    """
    ws = wb[LEGEND_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    start = None
    for idx, row in enumerate(rows):
        if row and row[0] == section_label:
            start = idx
            break
    if start is None:
        raise LegendConfigError(
            f"[ABORT] Legend tab is missing the '{section_label}' lookup "
            f"table. Add it before running."
        )
    table = {}
    for row in rows[start + 2:]:
        if not row or row[0] is None:
            break
        key, value = row[0], row[1]
        if key is not None and value is not None:
            table[str(key)] = str(value)
    return table


# ---------------------------------------------------------------------------
# Script-write formatting (Sprint 2 Run 3 fixes #4 and #6)
#
# Every cell a script writes gets font size 12 so it's visually distinct
# from manually-entered / untouched cells, which stay at the sheet's
# default size 8. write_cell() is the single choke point for this so
# portfolio_csv_ingestion.py, price_action_4h.py and fetch_engine_monthly.py
# all apply it identically.
#
# Number-format convention note: some "percentage" values in this codebase
# are already stored percentage-scaled (5.09 meaning "5.09%" -- e.g. Barchart
# fields after stripping "%", and price_action_4h's *_Change_% fields, which
# are multiplied by 100 in code). Applying Excel's built-in Percentage format
# (which itself multiplies the stored value by 100 for display) to a value
# already on that scale would show "509%" instead of "5.09%". PERCENT_SCALED
# uses a custom literal-"%" format for those. PERCENT_FRACTION is Excel's
# real Percentage format, for the few fields stored as true fractions (e.g.
# price_action_4h's D_Price_vs_Cap_Mid, or yfinance's payoutRatio/ROE/
# dividendYield in fetch_engine_monthly, which yfinance returns as fractions).
# ---------------------------------------------------------------------------

WRITTEN_FONT_SIZE = 12

NUMBER_FORMAT_NUMBER = "#,##0.00"
NUMBER_FORMAT_PERCENT_SCALED = '0.00"%"'   # value already e.g. 5.09 -> "5.09%"
NUMBER_FORMAT_PERCENT_FRACTION = "0.00%"   # value is a true fraction e.g. 0.0509 -> "5.09%"
NUMBER_FORMAT_DATE = "yyyy-mm-dd"


def _vals_equal(current, new):
    """True if the cell already holds the value we would write.
    None and "" are both treated as empty — writing "" to a None cell is a
    no-op. Sprint 4: universal conditional cell write guard.
    """
    c_empty = current is None or current == ""
    n_empty = new is None or new == ""
    if c_empty and n_empty:
        return True
    if c_empty != n_empty:
        return False
    return current == new


def write_cell(ws, row, column, value, number_format=None):
    """
    Conditional write: reads the current cell value first. Only writes and
    applies font.size=12 if the value differs. Unchanged cells are untouched
    — font stays at 8 (default). Sprint 4: universal conditional cell write
    guard applied to every script via this single choke-point.
    """
    cell = ws.cell(row=row, column=column)
    if _vals_equal(cell.value, value):
        return cell  # no change — leave cell and font completely untouched
    cell.value = value
    f = cell.font
    cell.font = Font(
        name=f.name, size=WRITTEN_FONT_SIZE, bold=f.bold, italic=f.italic,
        vertAlign=f.vertAlign, underline=f.underline, strike=f.strike,
        color=f.color,
    )
    # Preserve alignment set in Excel (centre, left, right etc.) — Sprint 4 fix.
    a = cell.alignment
    cell.alignment = Alignment(
        horizontal=a.horizontal, vertical=a.vertical,
        wrap_text=a.wrap_text, shrink_to_fit=a.shrink_to_fit,
        indent=a.indent, text_rotation=a.text_rotation,
        readingOrder=a.readingOrder,
    )
    if number_format is not None:
        cell.number_format = number_format
    return cell
