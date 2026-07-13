"""
PORTFOLIO / CSV Ingestion Script
Spec: 00_Portfolio_Automation_Spec_V5.5.md, Section 7e. Sprint 4.

Handles all file-based inputs from Cowork input folders. Checks each of the
6 defined filenames on every run, processes any present, skips any absent
independently (no all-or-nothing dependency).

    HL inputs      (Inputs/HL/):
        account-summary.csv     -> PORTFOLIO tab (clear-and-rewrite, qty>0 only)
        portfolio-summary.csv   -> PORTFOLIO_TRANSACTIONS tab (trade history)
        income-transactions.csv -> DIVIDENDS tab (dividend payments)

    Barchart inputs (Inputs/BC/):
        Custom.csv       -> Universe tab, S_BC_* (KR view)
        Performance.csv  -> Universe tab, S_BC_* (momentum)
        Fundamental.csv  -> Universe tab, S_BC_* (TTM/fundamentals)

Join keys: Code -> M_Ticker (HL) | M_BC_Ticker, falling back to M_Ticker
when M_BC_Ticker is blank (Barchart).

Processing order (Sprint 3/4):
  1. CSV -> PORTFOLIO (clear-and-rewrite, Wave 5)
  2. PORTFOLIO -> Universe S_ columns
  3. Ticker enrichment sweeps (contains-match, Wave 5)
  4. Barchart processing
  5. Gap logging to Scheduler

DIVIDENDS dedup key (Wave 5): Date + Type + Description + Amount + Ticker.
  Ticker enrichment runs before dedup check for incoming rows.
PORTFOLIO_TRANSACTIONS dedup: date + ticker + transaction type + amount.

Ticker enrichment (Wave 5): contains-match — Lookups description (needle)
contained within DIVIDENDS/TRANSACTIONS description (haystack).

Wave 4: S_Sector derive from Barchart removed. S_Sector authority is now
populate_sector_from_lookups() in fetch_engine_weekly.py.

Wave 2: Conditional cell write guard universal (via workbook_io.write_cell).
Wave 8: Structured PHASE/ACTION/TICKER/RUN_ID/UID logging.

Never touches: M_ columns, COVERAGE, D_ columns.
"""

import csv
import logging
import shutil
import sys
import uuid
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from workbook_io import (
    INPUTS_HL,
    INPUTS_BC,
    ARCHIVE_HL,
    ARCHIVE_BC,
    BASE_PATH,
    find_workbook,
    save_workbook_with_increment,
    write_cell,
    read_legend_scalars,
    NUMBER_FORMAT_NUMBER,
    NUMBER_FORMAT_PERCENT_SCALED,
    NUMBER_FORMAT_DATE,
)

# ---------------------------------------------------------------------------
# Logging (Wave 8)
# ---------------------------------------------------------------------------

LOG_DIR = BASE_PATH / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "portfolio_csv_ingestion.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("portfolio_csv_ingestion")

_ctx: dict = {}  # run-level context: run_id, uid
_VOCAB_KEYS = ("PHASE=", "ACTION=", "TICKER=", "RUN_ID=", "UID=")


def _log(level: str, phase: str, action: str, ticker: str, message: str) -> None:
    run_id = _ctx.get("run_id", "")
    uid    = _ctx.get("uid", "")
    line   = (
        f"PHASE={phase} ACTION={action} TICKER={ticker} "
        f"RUN_ID={run_id} UID={uid} | {message}"
    )
    for key in _VOCAB_KEYS:
        if key not in line:
            log.warning(f"[VOCAB_FAIL] Missing {key} in log line")
    getattr(log, level.lower())(line)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

HL_ENCODING = "cp1252"
BC_ENCODING = "utf-8-sig"

ACTION_ITEM_FIELDS = ["gap_type", "source_file", "ticker", "note", "detected_date", "due_date"]
SCHEDULER_SHEET   = "Scheduler"

NULL_VALUE_TOKENS = {"N/L", "N/A"}


def _is_bc_footer_row(symbol):
    if symbol is None:
        return False
    s = str(symbol).strip()
    return s.startswith("Downloaded from Barchart") or s.startswith("N/L -")


# Sprint 4 Wave 5: PORTFOLIO tab is now clear-and-rewrite.
# column_map maps CSV source column -> PORTFOLIO tab destination column.
# "Code" (ticker) is written explicitly as the key column.
HL_PORTFOLIO_FILE = "account-summary.csv"
HL_PORTFOLIO_CFG = {
    "sheet":      "PORTFOLIO",
    "source_key": "Code",
    "qty_col":    "Units held",   # filter: quantity > 0
    "column_map": {
        "Stock":         "Stock",
        "Price (pence)": "Price (pence)",
        "Value (£)":     "Value (£)",
        "Cost (£)":      "Cost (£)",
        "Gain/loss (£)": "Gain/loss (£)",
        "Gain/loss (%)": "Gain/loss (%)",
        "Units held":    "Units held",
    },
}

HL_APPEND_FILES = {
    "portfolio-summary.csv":  "PORTFOLIO_TRANSACTIONS",
    "income-transactions.csv": "DIVIDENDS",
}

BC_MATCH_FILES = {
    "Custom.csv": {
        "sheet":      "Universe",
        "source_key": "Symbol",
        "column_map": {
            "Sector":         "S_BC_Sector",
            "Market Cap":     "S_BC_MarketCap",
            "Beta":           "S_BC_Beta60M",
            "Div Yield(a)":   "S_BC_DivYieldFwd",
            "Div Payout%":    "S_BC_PayoutRatio",
            "5Y Div%":        "S_BC_DivGrowth5Y",
            "P/E fwd":        "S_BC_PE_Fwd",
            "ROE%":           "S_BC_ROE",
            "Debt/Equity":    "S_BC_DebtEquity",
            "Int Cov":        "S_BC_IntCoverage",
            "Analyst Rating": "S_BC_AnalystRating",
            "5Y Earn%":       "S_BC_EarningsGrowth5Y",
            "Dividend Date":  "S_BC_ExDivDate",
            "PEG":            "S_BC_PEG",
        },
    },
    "Performance.csv": {
        "sheet":      "Universe",
        "source_key": "Symbol",
        "column_map": {
            "Wtd Alpha": "S_BC_WtdAlpha",
            "5D %Chg":   "S_BC_Pct5D",
            "1M %Chg":   "S_BC_Pct1M",
            "3M %Chg":   "S_BC_Pct3M",
            "52W %Chg":  "S_BC_Pct52W",
            "YTD %Chg":  "S_BC_PctYTD",
        },
    },
    "Fundamental.csv": {
        "sheet":      "Universe",
        "source_key": "Symbol",
        "column_map": {
            "Market Cap":    "S_BC_MarketCap",
            "P/E ttm":       "S_BC_PE_TTM",
            "EPS ttm":       "S_BC_EPS_TTM",
            "Net Income(a)": "S_BC_NetIncome",
            "Beta":          "S_BC_Beta",
            "Dividend(a)":   "S_BC_DivAnnual",
            "Div Yield(a)":  "S_BC_DivYield",
            "Earnings Date": "S_BC_EarningsDate",
        },
        "fill_if_blank": {"S_BC_MarketCap"},
    },
}

# Sprint 4 Wave 4: column renames (_Pct suffix) applied in FIELD_TYPES.
# S_Sector removed — no longer derived from Barchart by this script.
FIELD_TYPES = {
    # BC Custom.csv
    "S_BC_Sector":           "text",
    "S_BC_MarketCap":        "number",
    "S_BC_Beta60M":          "number",
    "S_BC_DivYieldFwd":      "percent",
    "S_BC_PayoutRatio":      "percent",
    "S_BC_DivGrowth5Y":      "percent",
    "S_BC_PE_Fwd":           "number",
    "S_BC_ROE":              "percent",
    "S_BC_DebtEquity":       "number",
    "S_BC_IntCoverage":      "number",
    "S_BC_AnalystRating":    "number",
    "S_BC_EarningsGrowth5Y": "percent",
    "S_BC_ExDivDate":        "date",
    "S_BC_PEG":              "number",
    # BC Performance.csv
    "S_BC_WtdAlpha": "number",
    "S_BC_Pct5D":    "percent",
    "S_BC_Pct1M":    "percent",
    "S_BC_Pct3M":    "percent",
    "S_BC_Pct52W":   "percent",
    "S_BC_PctYTD":   "percent",
    # BC Fundamental.csv
    "S_BC_PE_TTM":       "number",
    "S_BC_EPS_TTM":      "number",
    "S_BC_NetIncome":    "number",
    "S_BC_Beta":         "number",
    "S_BC_DivAnnual":    "number",
    "S_BC_DivYield":     "percent",
    "S_BC_EarningsDate": "date",
    # HL account-summary.csv -> PORTFOLIO tab (column names identical)
    "Value (£)":     "number",
    "Cost (£)":      "number",
    "Units held":    "number",
    "Gain/loss (£)": "number",
    "Gain/loss (%)": "percent",
    "Price (pence)": "number",
    # Universe S_ columns (written by portfolio_to_universe_s_cols)
    "S_MarketValue_GBP": "number",
    "S_CostBasis":       "number",
    "S_UnitsHeld":       "number",
    "S_PnL_GBP":         "number",
    "S_PnL_Pct":         "percent",
    "S_Current_Price":   "number",
    "S_AvgCost":         "number",
}

_FIELD_TYPE_FORMATS = {
    "number":  NUMBER_FORMAT_NUMBER,
    "percent": NUMBER_FORMAT_PERCENT_SCALED,
    "date":    NUMBER_FORMAT_DATE,
    "text":    None,
}


def clean_field_value(raw, field_type):
    """Clean raw CSV string per field_type. NULL_VALUE_TOKENS -> None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if s == "":
        return None
    if s.upper() in NULL_VALUE_TOKENS:
        return None
    if field_type == "number":
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return raw
    if field_type == "percent":
        try:
            return float(s.replace(",", "").replace("%", ""))
        except ValueError:
            return raw
    if field_type == "date":
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            return raw
    return raw


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

def header_map(ws):
    """{header_name: column_index} from row 1."""
    return {
        cell.value: cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def key_row_map(ws, key_col_idx):
    """{key_value: row_index} for all data rows (row 2+)."""
    rows = {}
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=key_col_idx).value
        if val is not None:
            rows[str(val)] = row_idx
    return rows


def read_csv_rows(path, encoding):
    with open(path, newline="", encoding=encoding) as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Wave 5: PORTFOLIO clear-and-rewrite
# ---------------------------------------------------------------------------

def process_hl_portfolio_rewrite(wb, filename, cfg):
    """
    Sprint 4 Wave 5: clear PORTFOLIO tab (keep row-1 header) then rewrite
    from CSV. Include only rows where quantity (Units held) > 0.
    Source faithful — no M_Eliminated check at ingestion.
    """
    rows    = read_csv_rows(INPUTS_HL / filename, HL_ENCODING)
    ws      = wb[cfg["sheet"]]
    headers = header_map(ws)

    # Clear all data rows, keep header
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    qty_col_src = cfg["qty_col"]
    src_key     = cfg["source_key"]

    written = 0
    new_row = 2
    for row in rows:
        qty_raw = row.get(qty_col_src, "")
        try:
            qty = float(str(qty_raw).replace(",", "").strip()) if qty_raw else 0.0
        except ValueError:
            qty = 0.0
        if qty <= 0:
            continue

        ticker = row.get(src_key)
        if not ticker:
            continue

        # Write Code (key) column
        if src_key in headers:
            write_cell(ws, new_row, headers[src_key], str(ticker).strip())

        # Write mapped columns
        for src_col, dest_col in cfg["column_map"].items():
            if dest_col in headers and src_col in row:
                field_type = FIELD_TYPES.get(dest_col, "text")
                value = clean_field_value(row[src_col], field_type)
                write_cell(
                    ws, new_row, headers[dest_col], value,
                    number_format=_FIELD_TYPE_FORMATS.get(field_type),
                )
        new_row += 1
        written += 1

    _log("info", "INGEST", "PORTFOLIO_REWRITE", "",
         f"{written} rows written from CSV (qty>0 filter applied)")


# ---------------------------------------------------------------------------
# Sprint 3 Fix 3: PORTFOLIO -> Universe S_ column write
# ---------------------------------------------------------------------------

PORTFOLIO_TO_UNIVERSE_MAP = {
    "S_UnitsHeld":       "Units held",
    "S_CostBasis":       "Cost (£)",
    "S_MarketValue_GBP": "Value (£)",
    "S_PnL_GBP":         "Gain/loss (£)",
    "S_PnL_Pct":         "Gain/loss (%)",
}


def portfolio_to_universe_s_cols(wb):
    """Push 6 S_ holding columns from PORTFOLIO into Universe. S_AvgCost derived."""
    port_ws     = wb["PORTFOLIO"]
    uni_ws      = wb["Universe"]
    port_headers = header_map(port_ws)
    uni_headers  = header_map(uni_ws)

    code_col = port_headers.get("Code")
    if not code_col:
        _log("warning", "INGEST", "PORT_TO_UNI", "",
             "Code column not found in PORTFOLIO — skipped")
        return

    uni_ticker_col = uni_headers.get("M_Ticker")
    if not uni_ticker_col:
        _log("warning", "INGEST", "PORT_TO_UNI", "",
             "M_Ticker not found in Universe — skipped")
        return

    uni_row_map = key_row_map(uni_ws, uni_ticker_col)

    written = 0
    for port_row in range(2, port_ws.max_row + 1):
        ticker = port_ws.cell(row=port_row, column=code_col).value
        if not ticker or str(ticker).strip() == "":
            continue
        ticker = str(ticker).strip()
        if ticker not in uni_row_map:
            continue
        uni_row = uni_row_map[ticker]

        for s_col, port_col in PORTFOLIO_TO_UNIVERSE_MAP.items():
            src_col_idx  = port_headers.get(port_col)
            dest_col_idx = uni_headers.get(s_col)
            if not src_col_idx or not dest_col_idx:
                continue
            raw = port_ws.cell(row=port_row, column=src_col_idx).value
            field_type = FIELD_TYPES.get(s_col, "text")
            value = clean_field_value(str(raw) if raw is not None else None, field_type)
            write_cell(uni_ws, uni_row, dest_col_idx, value,
                       number_format=_FIELD_TYPE_FORMATS.get(field_type))
            written += 1

        # S_AvgCost = Cost (£) / Units held
        cost_idx  = port_headers.get("Cost (£)")
        units_idx = port_headers.get("Units held")
        avg_col   = uni_headers.get("S_AvgCost")
        if cost_idx and units_idx and avg_col:
            cost  = port_ws.cell(row=port_row, column=cost_idx).value
            units = port_ws.cell(row=port_row, column=units_idx).value
            try:
                avg = (
                    round(float(cost) / float(units), 6)
                    if (cost and units and float(units) != 0) else None
                )
            except (TypeError, ValueError):
                avg = None
            write_cell(uni_ws, uni_row, avg_col, avg,
                       number_format=_FIELD_TYPE_FORMATS.get("number"))
            if avg is not None:
                written += 1

    _log("info", "INGEST", "PORT_TO_UNI", "", f"{written} cell(s) written to Universe S_ columns")


# ---------------------------------------------------------------------------
# HL: append files (DIVIDENDS, PORTFOLIO_TRANSACTIONS)
# ---------------------------------------------------------------------------

def _normalize_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _normalize_amount(value):
    if value is None or value == "":
        return None
    try:
        return round(float(str(value).replace(",", "").strip()), 2)
    except ValueError:
        return None


def _normalize_ticker(value):
    return str(value or "").strip().upper()


def transaction_type_from_reference(reference):
    """B<digits>=BUY, S<digits>=SELL, anything else used as-is."""
    ref = str(reference or "").strip()
    if len(ref) > 1 and ref[0].upper() == "B" and ref[1:].isdigit():
        return "BUY"
    if len(ref) > 1 and ref[0].upper() == "S" and ref[1:].isdigit():
        return "SELL"
    return ref.upper()


def dividend_key(date_val, type_val, description_val, amount_val, ticker_val=""):
    """
    Sprint 4 Wave 5: dedup key = Date + Type + Description + Amount + Ticker.
    Ticker enrichment runs before dedup check for incoming rows so that
    incoming can match against already-enriched existing rows.
    """
    return (
        _normalize_date(date_val),
        str(type_val or "").strip().upper(),
        str(description_val or "").strip(),
        _normalize_amount(amount_val),
        _normalize_ticker(ticker_val),
    )


def transaction_key(date_val, ticker_val, reference_val, amount_val):
    return (
        _normalize_date(date_val),
        _normalize_ticker(ticker_val),
        transaction_type_from_reference(reference_val),
        _normalize_amount(amount_val),
    )


def existing_append_keys(ws, headers, sheet_name):
    """Composite keys already present in the sheet per dedup rule."""
    keys = set()
    if sheet_name == "DIVIDENDS":
        date_c   = headers.get("Date")
        type_c   = headers.get("Type")
        desc_c   = headers.get("Description")
        amount_c = headers.get("Amount")
        ticker_c = headers.get("Ticker")
        if not (date_c and amount_c):
            return keys
        for row_idx in range(2, ws.max_row + 1):
            d    = ws.cell(row=row_idx, column=date_c).value
            t    = ws.cell(row=row_idx, column=type_c).value   if type_c   else None
            desc = ws.cell(row=row_idx, column=desc_c).value   if desc_c   else None
            a    = ws.cell(row=row_idx, column=amount_c).value
            tk   = ws.cell(row=row_idx, column=ticker_c).value if ticker_c else None
            if d is None and a is None:
                continue
            keys.add(dividend_key(d, t, desc, a, tk or ""))
    elif sheet_name == "PORTFOLIO_TRANSACTIONS":
        date_c   = headers.get("Trade date")
        ticker_c = headers.get("Ticker")
        ref_c    = headers.get("Reference")
        amount_c = headers.get("Value (£)")
        if not (date_c and amount_c):
            return keys
        for row_idx in range(2, ws.max_row + 1):
            d = ws.cell(row=row_idx, column=date_c).value
            t = ws.cell(row=row_idx, column=ticker_c).value if ticker_c else None
            r = ws.cell(row=row_idx, column=ref_c).value    if ref_c    else None
            a = ws.cell(row=row_idx, column=amount_c).value
            if d is None and a is None:
                continue
            keys.add(transaction_key(d, t, r, a))
    return keys


def key_for_incoming_row(sheet_name, row, lookup_pairs=None):
    """
    Build dedup key for an incoming CSV row.
    For DIVIDENDS: attempts inline ticker enrichment via lookup_pairs before
    building the key, so incoming can match already-enriched existing rows.
    """
    if sheet_name == "DIVIDENDS":
        ticker = ""
        if lookup_pairs:
            ticker = match_ticker_by_description(
                row.get("Description"), lookup_pairs
            ) or ""
        return dividend_key(
            row.get("Date"), row.get("Type"),
            row.get("Description"), row.get("Amount"), ticker,
        )
    if sheet_name == "PORTFOLIO_TRANSACTIONS":
        return transaction_key(
            row.get("Trade date"), row.get("Ticker"),
            row.get("Reference"), row.get("Value (£)"),
        )
    return None


def process_hl_append_file(wb, filename, sheet_name, lookup_pairs=None):
    rows    = read_csv_rows(INPUTS_HL / filename, HL_ENCODING)
    ws      = wb[sheet_name]
    headers = header_map(ws)

    existing_keys = existing_append_keys(ws, headers, sheet_name)

    next_row = ws.max_row + 1
    appended, skipped = 0, 0
    for row in rows:
        key = key_for_incoming_row(sheet_name, row, lookup_pairs)
        if key is not None and key in existing_keys:
            skipped += 1
            continue

        for col_name, value in row.items():
            if col_name in headers:
                write_cell(ws, next_row, headers[col_name], value)

        if key is not None:
            existing_keys.add(key)
        next_row += 1
        appended += 1

    _log("info", "INGEST", f"APPEND_{sheet_name}", "",
         f"{appended} appended, {skipped} dupes skipped")
    return appended, skipped


# ---------------------------------------------------------------------------
# Ticker enrichment (contains-match, Sprint 4 Wave 5) + DIVIDENDS Month
# ---------------------------------------------------------------------------

def open_position_tickers(wb):
    """Set of ticker codes in PORTFOLIO tab (open positions only)."""
    ws = wb["PORTFOLIO"]
    headers  = header_map(ws)
    code_col = headers.get("Code")
    if not code_col:
        return set()
    tickers = set()
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=code_col).value
        if val:
            tickers.add(str(val).strip().upper())
    return tickers


LOOKUPS_SHEET = "Lookups"


def build_lookup_pairs(wb):
    """
    [(ticker, description), ...] from Lookups 'Ticker'/'Description' columns,
    resolved by header name. Sorted longest-first so the most specific match wins.
    Sprint 3 Fix 4: header-name resolution.
    """
    ws   = wb[LOOKUPS_SHEET]
    hmap = header_map(ws)
    ticker_col = hmap.get("Ticker")
    desc_col   = hmap.get("Description")
    if not ticker_col or not desc_col:
        raise KeyError(
            f"Lookups tab missing 'Ticker' or 'Description' header "
            f"(found: {list(hmap.keys())})"
        )
    pairs = []
    for row_idx in range(2, ws.max_row + 1):
        ticker      = ws.cell(row=row_idx, column=ticker_col).value
        description = ws.cell(row=row_idx, column=desc_col).value
        if ticker and description:
            pairs.append((str(ticker), str(description)))
    pairs.sort(key=lambda p: -len(p[1]))
    return pairs


def match_ticker_by_description(description, lookup_pairs):
    """
    Sprint 4 Wave 5: contains-match. Lookups description (needle) must be
    contained within the DIVIDENDS/TRANSACTIONS description (haystack).
    Longest entry checked first — most specific match wins.
    Previously used startswith; changed to handle HL verbose strings.
    """
    if not description:
        return None
    for ticker, lookup_desc in lookup_pairs:
        if lookup_desc in description:
            return ticker
    return None


def _month_str(date_val):
    """DIVIDENDS Month column — 'YYYY-MM'."""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m")
    s = str(date_val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return None


def enrich_dividends(wb, lookup_pairs, run_date, open_tickers=None):
    """Sweep DIVIDENDS: fill blank Ticker (contains-match) and blank Month."""
    ws      = wb["DIVIDENDS"]
    headers = header_map(ws)
    date_c, desc_c, ticker_c, month_c = (
        headers.get("Date"), headers.get("Description"),
        headers.get("Ticker"), headers.get("Month"),
    )
    matched, gapped, month_filled = 0, 0, 0
    gaps = []
    for row_idx in range(2, ws.max_row + 1):
        description = ws.cell(row=row_idx, column=desc_c).value if desc_c else None

        if ticker_c:
            existing_ticker = ws.cell(row=row_idx, column=ticker_c).value
            if existing_ticker is not None and str(existing_ticker).strip() != "":
                # Already has ticker — skip enrichment, check closed-position
                if open_tickers is not None:
                    pass  # no action; month fill still runs below
            else:
                found = match_ticker_by_description(description, lookup_pairs)
                if found:
                    if open_tickers is None or found.upper() in open_tickers:
                        write_cell(ws, row_idx, ticker_c, found)
                        matched += 1
                        _log("info", "ENRICH", "DIV_TICKER", found, f"row {row_idx}")
                    # else: closed position — skip silently
                elif description:
                    gapped += 1
                    gaps.append(description)

        if month_c and date_c:
            existing_month = ws.cell(row=row_idx, column=month_c).value
            if existing_month is None or str(existing_month).strip() == "":
                date_val  = ws.cell(row=row_idx, column=date_c).value
                month_val = _month_str(date_val)
                if month_val:
                    write_cell(ws, row_idx, month_c, month_val)
                    month_filled += 1

    gap_records = [
        {
            "gap_type":      "DIVIDENDS Ticker Gap",
            "source_file":   "DIVIDENDS",
            "ticker":        "",
            "note":          f"DIVIDENDS ticker gap — {desc} — manual Lookups update required.",
            "detected_date": run_date.date(),
        }
        for desc in gaps
    ]
    _log("info", "ENRICH", "DIV_SUMMARY", "",
         f"{matched} matched, {gapped} gap(s), {month_filled} Month filled")
    return matched, gapped, month_filled, gap_records


def enrich_portfolio_transactions(wb, lookup_pairs, run_date, open_tickers=None):
    """
    Sweep PORTFOLIO_TRANSACTIONS: fill blank Ticker via contains-match.
    BUY/SELL rows only — non-trade cash movements left blank (no gap logged).
    Sprint 4 Wave 5: contains-match replaces starts-with.
    """
    ws      = wb["PORTFOLIO_TRANSACTIONS"]
    headers = header_map(ws)
    desc_c, ticker_c, ref_c = (
        headers.get("Description"), headers.get("Ticker"), headers.get("Reference"),
    )
    matched, gapped = 0, 0
    gaps = []
    for row_idx in range(2, ws.max_row + 1):
        if ticker_c is None:
            break
        existing_ticker = ws.cell(row=row_idx, column=ticker_c).value
        if existing_ticker is not None and str(existing_ticker).strip() != "":
            continue

        reference = ws.cell(row=row_idx, column=ref_c).value if ref_c else None
        if transaction_type_from_reference(reference) not in ("BUY", "SELL"):
            continue  # cash movement — legitimately no ticker

        description = ws.cell(row=row_idx, column=desc_c).value if desc_c else None
        found = match_ticker_by_description(description, lookup_pairs)
        if found:
            write_cell(ws, row_idx, ticker_c, found)
            matched += 1
            _log("info", "ENRICH", "TXN_TICKER", found, f"row {row_idx}")
            # closed position: write but don't gap-log
        elif description:
            gapped += 1
            gaps.append(description)

    gap_records = [
        {
            "gap_type":      "PORTFOLIO_TRANSACTIONS Ticker Gap",
            "source_file":   "PORTFOLIO_TRANSACTIONS",
            "ticker":        "",
            "note":          f"PORTFOLIO_TRANSACTIONS ticker gap — {desc} — manual Lookups update required.",
            "detected_date": run_date.date(),
        }
        for desc in gaps
    ]
    _log("info", "ENRICH", "TXN_SUMMARY", "", f"{matched} matched, {gapped} gap(s)")
    return matched, gapped, gap_records


# ---------------------------------------------------------------------------
# Barchart: match-and-update
# ---------------------------------------------------------------------------

def effective_bc_ticker(headers, ws, row_idx):
    """M_BC_Ticker if populated, else fall back to M_Ticker."""
    bc_col = headers.get("M_BC_Ticker")
    val = ws.cell(row=row_idx, column=bc_col).value if bc_col else None
    if val:
        return str(val)
    m_col = headers.get("M_Ticker")
    val   = ws.cell(row=row_idx, column=m_col).value if m_col else None
    return str(val) if val else None


def universe_ticker_row_map(ws):
    """{effective_ticker: row_index} across Universe tab."""
    headers = header_map(ws)
    mapping = {}
    for row_idx in range(2, ws.max_row + 1):
        ticker = effective_bc_ticker(headers, ws, row_idx)
        if ticker:
            mapping[ticker] = row_idx
    return mapping


def process_bc_match_file(wb, filename, cfg, ticker_row_map):
    """
    Write Barchart CSV values to Universe tab.
    Sprint 4 Wave 4: S_Sector derive block REMOVED. S_BC_Sector is written
    faithfully; S_Sector authority moved to populate_sector_from_lookups()
    in fetch_engine_weekly.py.
    """
    rows    = read_csv_rows(INPUTS_BC / filename, BC_ENCODING)
    ws      = wb[cfg["sheet"]]
    headers = header_map(ws)
    fill_if_blank = cfg.get("fill_if_blank", set())

    written = 0
    for row in rows:
        ticker = row.get(cfg["source_key"])
        if not ticker or _is_bc_footer_row(ticker) or ticker not in ticker_row_map:
            continue
        row_idx = ticker_row_map[ticker]
        for src_col, dest_col in cfg["column_map"].items():
            if dest_col in headers and src_col in row:
                if dest_col in fill_if_blank:
                    current = ws.cell(row=row_idx, column=headers[dest_col]).value
                    if current is not None and str(current).strip() != "":
                        continue
                field_type = FIELD_TYPES.get(dest_col, "text")
                value      = clean_field_value(row[src_col], field_type)
                write_cell(
                    ws, row_idx, headers[dest_col], value,
                    number_format=_FIELD_TYPE_FORMATS.get(field_type),
                )
                written += 1
                # NOTE: S_Sector is NOT derived here. It is authoritative from
                # Lookups tab via fetch_engine_weekly.populate_sector_from_lookups().

    _log("info", "INGEST", f"BC_{filename}", "", f"{written} cells written")


# ---------------------------------------------------------------------------
# Archiving
# ---------------------------------------------------------------------------

def archive_file(src_path, archive_dir, timestamp):
    archive_dir.mkdir(parents=True, exist_ok=True)
    stamped_name = f"{src_path.stem}_{timestamp}{src_path.suffix}"
    shutil.move(str(src_path), str(archive_dir / stamped_name))


# ---------------------------------------------------------------------------
# Gap detection & Scheduler action items
# ---------------------------------------------------------------------------

def expected_watchlist_tickers(ws):
    headers      = header_map(ws)
    watchlist_col = headers.get("M_BC_Watchlist")
    if not watchlist_col:
        return set()
    expected = set()
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=watchlist_col).value == "Y":
            ticker = effective_bc_ticker(headers, ws, row_idx)
            if ticker:
                expected.add(ticker)
    return expected


def detect_gaps(present_files, expected_tickers, run_date, gap_due_days):
    gaps          = []
    detected_date = run_date.date() if hasattr(run_date, "date") else run_date
    due_date      = detected_date + timedelta(days=gap_due_days)
    for filename, found_tickers in present_files.items():
        if found_tickers is None:
            gaps.append({
                "gap_type": "Type A", "source_file": filename, "ticker": "",
                "note": "expected file missing for this run",
                "detected_date": detected_date, "due_date": due_date,
            })
        else:
            for ticker in sorted(expected_tickers - found_tickers):
                gaps.append({
                    "gap_type": "Type B", "source_file": filename, "ticker": ticker,
                    "note": "expected ticker missing from file",
                    "detected_date": detected_date, "due_date": due_date,
                })
    return gaps


def find_action_item_header_row(ws):
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == ACTION_ITEM_FIELDS[0]:
            return row_idx
    return None


def log_gaps_to_scheduler(wb, gaps, gap_due_days):
    if not gaps:
        return
    ws         = wb[SCHEDULER_SHEET]
    header_row = find_action_item_header_row(ws)
    if header_row is None:
        header_row = ws.max_row + 2
        for col_idx, name in enumerate(ACTION_ITEM_FIELDS, start=1):
            write_cell(ws, header_row, col_idx, name)

    next_row = ws.max_row + 1
    for gap in gaps:
        detected_date = gap["detected_date"]
        due_date      = gap.get("due_date") or (detected_date + timedelta(days=gap_due_days))
        row_values    = dict(gap, due_date=due_date)
        for col_idx, field in enumerate(ACTION_ITEM_FIELDS, start=1):
            value = row_values[field]
            fmt   = NUMBER_FORMAT_DATE if field in ("detected_date", "due_date") else None
            write_cell(ws, next_row, col_idx, value, number_format=fmt)
        next_row += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(workbook_path=None):
    _ctx["run_id"] = str(uuid.uuid4())[:8]
    _ctx["uid"]    = ""  # ICU user sessions not yet wired

    run_date  = datetime.now()
    timestamp = run_date.strftime("%Y%m%d_%H%M%S")

    _log("info", "STARTUP", "RUN_START", "",
         f"portfolio_csv_ingestion starting — wb={workbook_path or 'auto'}")

    resolved_by_glob = workbook_path is None
    workbook_path    = find_workbook() if resolved_by_glob else Path(workbook_path)

    wb = openpyxl.load_workbook(
        workbook_path, keep_vba=workbook_path.suffix == ".xlsm"
    )

    legend_scalars = read_legend_scalars(wb, ["GAP_DUE_DAYS"])
    gap_due_days   = int(legend_scalars["GAP_DUE_DAYS"])

    # Build lookup_pairs early — used for DIVIDENDS dedup pre-enrichment
    try:
        lookup_pairs = build_lookup_pairs(wb)
        _log("info", "STARTUP", "LOOKUPS_LOAD", "",
             f"{len(lookup_pairs)} Lookups pairs loaded")
    except KeyError as exc:
        _log("error", "STARTUP", "LOOKUPS_LOAD", "", str(exc))
        lookup_pairs = []

    # --- Step 1: PORTFOLIO clear-and-rewrite (Wave 5) ---
    portfolio_path = INPUTS_HL / HL_PORTFOLIO_FILE
    if portfolio_path.exists():
        _log("info", "INGEST", "PORTFOLIO_START", "",
             f"Processing {HL_PORTFOLIO_FILE}")
        try:
            process_hl_portfolio_rewrite(wb, HL_PORTFOLIO_FILE, HL_PORTFOLIO_CFG)
            archive_file(portfolio_path, ARCHIVE_HL, timestamp)
        except Exception as exc:
            _log("error", "INGEST", "PORTFOLIO_ERR", "", str(exc))

    # --- Step 1b: HL append files ---
    for filename, sheet_name in HL_APPEND_FILES.items():
        path = INPUTS_HL / filename
        if path.exists():
            _log("info", "INGEST", "APPEND_START", "",
                 f"Processing {filename} -> {sheet_name}")
            try:
                lp = lookup_pairs if sheet_name == "DIVIDENDS" else None
                process_hl_append_file(wb, filename, sheet_name, lookup_pairs=lp)
                archive_file(path, ARCHIVE_HL, timestamp)
            except Exception as exc:
                _log("error", "INGEST", "APPEND_ERR", "", f"{filename}: {exc}")

    # --- Step 2: PORTFOLIO -> Universe S_ columns ---
    _log("info", "INGEST", "PORT_TO_UNI_START", "", "Pushing S_ columns to Universe")
    try:
        portfolio_to_universe_s_cols(wb)
    except Exception as exc:
        _log("error", "INGEST", "PORT_TO_UNI_ERR", "", str(exc))

    # --- Step 3: Ticker enrichment + Month ---
    _log("info", "ENRICH", "ENRICH_START", "", "Running ticker enrichment sweeps")
    open_tickers = open_position_tickers(wb)

    div_matched, div_gapped, month_filled, div_gap_records = enrich_dividends(
        wb, lookup_pairs, run_date, open_tickers=open_tickers
    )
    pt_matched, pt_gapped, pt_gap_records = enrich_portfolio_transactions(
        wb, lookup_pairs, run_date, open_tickers=open_tickers
    )

    # --- Step 4: Barchart inputs ---
    _log("info", "INGEST", "BC_START", "", "Processing Barchart inputs")
    universe_ws      = wb["Universe"]
    ticker_row_map   = universe_ticker_row_map(universe_ws)
    expected_tickers = expected_watchlist_tickers(universe_ws)

    present_files = {}
    for filename, cfg in BC_MATCH_FILES.items():
        path = INPUTS_BC / filename
        if path.exists():
            try:
                rows = read_csv_rows(path, BC_ENCODING)
                found_tickers = {
                    r[cfg["source_key"]] for r in rows
                    if r.get(cfg["source_key"]) and not _is_bc_footer_row(r[cfg["source_key"]])
                }
                process_bc_match_file(wb, filename, cfg, ticker_row_map)
                archive_file(path, ARCHIVE_BC, timestamp)
                present_files[filename] = found_tickers
            except Exception as exc:
                _log("error", "INGEST", "BC_ERR", "", f"{filename}: {exc}")
                present_files[filename] = None
        else:
            present_files[filename] = None

    # --- Step 5: Gap logging ---
    gaps     = detect_gaps(present_files, expected_tickers, run_date, gap_due_days)
    all_gaps = gaps + div_gap_records + pt_gap_records
    log_gaps_to_scheduler(wb, all_gaps, gap_due_days)

    if resolved_by_glob:
        workbook_path = save_workbook_with_increment(wb, workbook_path)
    else:
        wb.save(workbook_path)

    _log("info", "COMPLETE", "RUN_END", "",
         f"{len(gaps)} BC gap(s), {div_gapped + pt_gapped} ticker gap(s). "
         f"Workbook: {workbook_path.name}")
    return wb


if __name__ == "__main__":
    run()
