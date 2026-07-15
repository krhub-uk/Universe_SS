"""
Price Action Script -- Intraday cadence
Spec: 00_Portfolio_Automation_Spec_V5.5.md, Section 7a. Sprint 4.
Originally: Sprint 2 script re-engineering, replaces 1H/4H scope of price_action_4h.py.

Runs: 1H/4H cadence (11am / 3pm / 7pm, or hourly -- schedule is external to
this script; every run behaves identically).

Scope each run:
    M_Eliminated != "No Touch"   (universal kill switch, checked first)
    -- all rows, all sleeves. No M_Fetch_Cadence filter (cadence is
    now governed by the run schedule, not a per-row column check).

S_ columns fed (9 attributes only):
    S_Daily_High, S_Daily_Low, S_Daily_Open, S_Last_Price, S_Current_Price,
    S_LastTradedTime, S_1D_Change_Pct, S_2D_Change_Pct, S_3D_Change_Pct

No D_ columns. No CSV ingestion. No cadence filter.

Source: one yfinance Ticker.history() call per row, no .info[] calls.

Notes on scope decisions not fully specified in §7a itself:
  - §7a's literal text says "history(1d) only" as the lightest possible YF
    call. A single history(period="1d") pull returns exactly one daily bar,
    which cannot support S_1D/2D/3D_Change_Pct (each needs a prior close N
    trading days back). Raised with the user directly -- resolved as: pull
    history(period="5d") instead (still one Ticker call, still zero .info[]
    calls) so the three change-% columns actually compute.
  - S_Current_Price / S_Last_Price: both set from the latest daily bar's
    Close -- true intraday tick not available without .info[].
  - S_LastTradedTime: uses the timestamp of the latest returned bar
    (hist.index[-1]) formatted "YYYY-MM-DD HH:MM:SS".
  - Column names S_1D/2D/3D_Change_Pct: Sprint 4 Wave 4 rename from _%.

Wave 8: Structured PHASE/ACTION/TICKER/RUN_ID/UID logging.

openpyxl constraint: surgical cell edits only, never a pandas rewrite.
"""

import logging
import sys
import uuid
import numpy as np
import yfinance as yf
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

from workbook_io import (
    BASE_PATH,
    find_workbook,
    save_workbook_with_increment,
    write_cell,
    NUMBER_FORMAT_NUMBER,
    NUMBER_FORMAT_PERCENT_SCALED,
)
from icu_client import push_status, check_gate, resolve_version

COMPONENT_ID = "universe_ss_intraday"

# ---------------------------------------------------------------------------
# Logging (Wave 8)
# ---------------------------------------------------------------------------

LOG_DIR = BASE_PATH / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "price_action_intraday.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("price_action_intraday")

_ctx: dict = {}
_VOCAB_KEYS = ("PHASE=", "ACTION=", "TICKER=", "RUN_ID=", "UID=")


def _log(level: str, phase: str, action: str, ticker: str, message: str) -> None:
    run_id = _ctx.get("run_id", "")
    uid    = _ctx.get("uid", "")
    line   = (
        f"PHASE={phase} ACTION={action} TICKER={ticker} "
        f"RUN_ID={run_id} UID={uid} | {message}"
    )
    for key in _VOCAB_KEYS:
        if key not in line:
            log.warning(f"[VOCAB_FAIL] Missing {key}")
    getattr(log, level.lower())(line)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEET_NAME = "Universe"
NO_TOUCH   = "No Touch"

S_COLUMNS_INTRADAY = [
    "S_Daily_High", "S_Daily_Low", "S_Daily_Open",
    "S_Last_Price", "S_Current_Price", "S_LastTradedTime",
    "S_1D_Change_Pct", "S_2D_Change_Pct", "S_3D_Change_Pct",  # Wave 4 renames
]

FIELD_TYPES = {
    "S_Daily_High":     "number",
    "S_Daily_Low":      "number",
    "S_Daily_Open":     "number",
    "S_Last_Price":     "number",
    "S_Current_Price":  "number",
    "S_LastTradedTime": "text",
    "S_1D_Change_Pct":  "percent_scaled",
    "S_2D_Change_Pct":  "percent_scaled",
    "S_3D_Change_Pct":  "percent_scaled",
}

_FIELD_TYPE_FORMATS = {
    "number":         NUMBER_FORMAT_NUMBER,
    "percent_scaled": NUMBER_FORMAT_PERCENT_SCALED,
    "text":           None,
}


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

def header_map(ws):
    """{header_name: column_index} from row 1."""
    return {cell.value: cell.column for cell in ws[1] if cell.value is not None}


def get_cell(ws, headers, row_idx, col_name, default=None):
    col = headers.get(col_name)
    if col is None:
        return default
    val = ws.cell(row=row_idx, column=col).value
    return val if val is not None else default


def set_cell(ws, headers, row_idx, col_name, value):
    col = headers.get(col_name)
    if col is None:
        return  # column not in sheet -- skip silently
    field_type = FIELD_TYPES.get(col_name, "text")
    write_cell(ws, row_idx, col, value, number_format=_FIELD_TYPE_FORMATS.get(field_type))


# ---------------------------------------------------------------------------
# Scope filter
# ---------------------------------------------------------------------------

def in_scope(ws, headers, row_idx):
    """M_Eliminated != "No Touch" -- all rows, all sleeves, no cadence filter."""
    return get_cell(ws, headers, row_idx, "M_Eliminated") != NO_TOUCH


def resolve_ticker(ws, headers, row_idx):
    yf_ticker = get_cell(ws, headers, row_idx, "S_YF_Ticker")
    if yf_ticker:
        return str(yf_ticker)
    m_ticker = get_cell(ws, headers, row_idx, "M_Ticker")
    return str(m_ticker) if m_ticker else None


# ---------------------------------------------------------------------------
# yfinance fetch
# ---------------------------------------------------------------------------

def fetch_price_action(symbol):
    """
    One yfinance Ticker object per row, one history() call, no .info[].
    period="5d" so S_1D/2D/3D_Change_Pct can be computed from prior closes.
    Returns dict of 9 S_ columns, or None on failure.
    """
    try:
        t    = yf.Ticker(symbol)
        hist = t.history(period="5d")
    except Exception as exc:
        _log("warning", "FETCH", "YF_ERROR", symbol, f"yfinance error: {exc}")
        return None

    if hist.empty:
        _log("warning", "FETCH", "YF_EMPTY", symbol, "empty history")
        return None

    last       = hist.iloc[-1]
    closes     = hist["Close"]
    last_price = float(last["Close"])

    def pct_change(n):
        if len(closes) <= n:
            return None
        prior = closes.iloc[-1 - n]
        if prior in (0, None) or (isinstance(prior, float) and np.isnan(prior)):
            return None
        return (last_price - prior) / prior * 100

    return {
        "S_Daily_High":    float(last["High"]),
        "S_Daily_Low":     float(last["Low"]),
        "S_Daily_Open":    float(last["Open"]),
        "S_Last_Price":    last_price,
        "S_Current_Price": last_price,
        "S_LastTradedTime": hist.index[-1].strftime("%Y-%m-%d %H:%M:%S"),
        "S_1D_Change_Pct": pct_change(1),
        "S_2D_Change_Pct": pct_change(2),
        "S_3D_Change_Pct": pct_change(3),
    }


# ---------------------------------------------------------------------------
# Main run
# ---------------------------------------------------------------------------

def run(workbook_path=None):
    _ctx["run_id"] = str(uuid.uuid4())[:8]
    _ctx["uid"]    = ""

    start = datetime.now()
    _log("info", "STARTUP", "RUN_START", "", "price_action_intraday.py starting")

    resolved_by_glob = workbook_path is None
    workbook_path    = find_workbook() if resolved_by_glob else Path(workbook_path)

    wb      = openpyxl.load_workbook(workbook_path, keep_vba=str(workbook_path).endswith(".xlsm"))
    ws      = wb[SHEET_NAME]
    headers = header_map(ws)

    processed, skipped = 0, 0
    for row_idx in range(2, ws.max_row + 1):
        if not in_scope(ws, headers, row_idx):
            continue

        symbol = resolve_ticker(ws, headers, row_idx)
        if not symbol:
            skipped += 1
            continue

        s_values = fetch_price_action(symbol)
        if s_values is None:
            skipped += 1
            continue

        for col_name in S_COLUMNS_INTRADAY:
            set_cell(ws, headers, row_idx, col_name, s_values[col_name])

        _log("info", "FETCH", "TICKER_DONE", symbol, f"row {row_idx} written")
        processed += 1

    _ctx["processed"] = processed
    _ctx["skipped"] = skipped

    if resolved_by_glob:
        workbook_path = save_workbook_with_increment(wb, workbook_path)
    else:
        wb.save(workbook_path)

    elapsed = str(datetime.now() - start).split(".")[0]
    _log("info", "COMPLETE", "RUN_END", "",
         f"{processed} row(s) updated, {skipped} skipped. Duration {elapsed}")
    return wb


if __name__ == "__main__":
    workbook_path = find_workbook()
    VERSION = resolve_version(str(workbook_path))

    if not check_gate(COMPONENT_ID):
        push_status(
            COMPONENT_ID, "PAUSED", VERSION,
            last_run_result="SKIPPED",
            trigger="GATE_CHECK"
        )
        sys.exit(0)

    run_start = datetime.now(timezone.utc).isoformat()
    push_status(
        COMPONENT_ID, "RUNNING", VERSION,
        last_run_utc=run_start,
        trigger="SCHEDULED"
    )

    try:
        run()
        processed = _ctx.get("processed", 0)
        skipped = _ctx.get("skipped", 0)
        if skipped == 0:
            result = "SUCCESS"
        elif processed > 0:
            result = "PARTIAL"
        else:
            result = "DEGRADED"
        push_status(
            COMPONENT_ID, "IDLE", VERSION,
            last_run_utc=run_start,
            last_run_result=result,
            trigger="SCHEDULED",
            metrics={"rows_processed": processed}
        )
    except Exception as e:
        push_status(
            COMPONENT_ID, "ERROR", VERSION,
            last_run_utc=run_start,
            last_run_result="FAILED",
            message=str(e)
        )
        raise
