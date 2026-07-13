"""
derive_engine.py — D_ computation pass
Spec: 00_Portfolio_Automation_Spec_V5.5.md, Section 7e. Sprint 4.

Single owner of all D_ column computation. Reads S_ and M_ values from
Universe, computes all D_ columns, writes results back via openpyxl surgical
cell writes. NO Excel formulas — immune to #REF! errors.

Cadence: runs after fetch_engine_weekly.py (via run_weekly.sh).
Bolt-on pattern: new derived metrics added as functions registered in BOLT_ONS.

Execution order:
  1. DIVIDENDS aggregates (D_All_Income, D_AnnualIncome, D_Income_Payback_Pct)
  1b. QScore 3-column shift chain (read existing S_QScore, prep for comparison)
  2. QScore v2 (4 equal-weight inputs; 5th Momentum input if data available)
  2b. QScore shift chain write (S_QScore / S_QScore_Prior updated on change)
  3. Price derived (D_PE_Discount, D_Yield_Premium_Pct, D_Price_vs_52W_High)
  4. Sell Board signals (Curator -> Grower -> Chartist -> Treasurer -> Comptroller)
  5. Income / YOC (D_Sleeve_Weight, D_YOC, D_YoC_PerHolding, D_GBP_Per_GBP1_Div)
  6. Bolt-ons (D_ValueClass etc)

Sprint 4 changes:
  Wave 3: D_Income_Payback_Pct stored x100 (raw %). QScore 3-column shift chain.
           M_Asset_Type eligibility gate added. Payout scoring bands updated for
           S_PayoutRatio_Pct (stored as 64.0 not 0.64).
  Wave 4: All _Pct column renames applied (S_DivGrowth_5Y_Pct, S_PayoutRatio_Pct,
           S_Dividend_Yield_Pct, D_Yield_Premium_Pct etc.).
  Wave 6: QScore Momentum bucket (5th input): S_Price_5Y_Return_Pct x0.50 +
           S_BC_Pct52W x0.50. Blank if either input blank. Bands TBD (calibrate
           after first real data run).
  Wave 8: Structured PHASE/ACTION/TICKER/RUN_ID/UID logging.

D_ columns written by price_action_eod.py (EOD cadence, not touched here):
  D_Cap_Candle_Mid, D_Price_vs_Cap_Mid, D_Close_Direction, D_Volume_Delta_Pct,
  D_Volume_Flag, D_Context_Flag, D_Cap_Mid_Anchor, D_Cap_Anchor_Date,
  D_Cap_Anchor_Active
"""

import logging
import sys
import uuid
from datetime import datetime, date
from pathlib import Path

import openpyxl
from dateutil.relativedelta import relativedelta

from workbook_io import BASE_PATH, find_workbook, write_cell, save_workbook_with_increment

# ---------------------------------------------------------------------------
# Logging (Wave 8)
# ---------------------------------------------------------------------------

LOG_DIR = BASE_PATH / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "derive_engine.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("derive_engine")

_ctx: dict = {}
_VOCAB_KEYS = ("PHASE=", "ACTION=", "TICKER=", "RUN_ID=", "UID=")


def _log(level: str, phase: str, action: str, ticker: str, message: str) -> None:
    run_id = _ctx.get("run_id", "")
    uid    = _ctx.get("uid", "")
    line   = (
        f"PHASE={phase} ACTION={action} TICKER={ticker} "
        f"RUN_ID={run_id} UID={uid} | {message}"
    )
    for key in _VOCAB_KEYS:
        if key not in line:
            log.warning(f"[VOCAB_FAIL] Missing {key}")
    getattr(log, level.lower())(line)


# ---------------------------------------------------------------------------
# QScore v2 — eligibility and scoring
# ---------------------------------------------------------------------------

QSCORE_V2_ELIGIBLE_CLASSES = {
    "Aristocrat_King", "Aristocrat", "Achiever", "Contender",
    "HighIncome", "MedIncome", "LowNoIncome",
}

SLEEVE_STREAK_TIERS = {"Aristocrat_King", "Aristocrat", "Achiever", "Contender"}
ELIGIBLE_CLASSES    = QSCORE_V2_ELIGIBLE_CLASSES  # alias for Sell Board

STREAK_PROXY_V2 = {
    "Aristocrat_King": 100,
    "Aristocrat":       80,
    "Achiever":         60,
    "Contender":        50,
    "HighIncome":       40,
    "MedIncome":        30,
    "LowNoIncome":      20,
}


def _score_divgrowth_v2(v):
    """DivGrowth_5Y_Pct banding (value stored as x100, e.g. 5.0 = 5%)."""
    if v >= 10.0: return 100
    if v >= 5.0:  return 80
    if v >= 2.0:  return 60
    if v >= 0.0:  return 40
    return 0


def _score_payout_v2(v):
    """PayoutRatio_Pct banding (Sprint 4: value stored x100, e.g. 64.0 = 64%).
    Bands: <=40=100, 40-60=80, 60-75=60, 75-90=40, >90=0.
    Previously used fraction scale (0.40 etc.) -- updated for _Pct rename."""
    if v <= 40.0: return 100
    if v <= 60.0: return 80
    if v <= 75.0: return 60
    if v <= 90.0: return 40
    return 0


def _score_beta_v2(v):
    """Beta banding: 0.50-0.75=100, 0.75-1.00=80, 1.00-1.25=60, 0.25-0.50=40, else=0."""
    if 0.50 <= v <= 0.75: return 100
    if 0.75 <  v <= 1.00: return 80
    if 1.00 <  v <= 1.25: return 60
    if 0.25 <= v <  0.50: return 40
    return 0


def _score_momentum_v2(price_5y_pct, bc_pct52w):
    """
    Wave 6: QScore Momentum bucket (5th input).
    Composite = S_Price_5Y_Return_Pct x 0.50 + S_BC_Pct52W x 0.50.
    Returns None if either input is blank.
    Scoring bands TBD -- calibrate after first real data run.
    Using placeholder linear scale: composite >= 30 = 100, >= 10 = 80,
    >= 0 = 60, >= -10 = 40, < -10 = 0. Update after calibration.
    """
    if price_5y_pct is None or bc_pct52w is None:
        return None
    composite = float(price_5y_pct) * 0.50 + float(bc_pct52w) * 0.50
    # Placeholder bands -- calibrate on first real data run
    if composite >= 30: return 100
    if composite >= 10: return 80
    if composite >= 0:  return 60
    if composite >= -10: return 40
    return 0


def compute_qscore(row, cm):
    """
    QScore v2: 4 equal-weight inputs averaged (+ 5th Momentum if available).
    Any blank core input -> D_QTier = 'Insufficient Data', D_QScore = blank.
    Eligibility: M_Eliminated != No Touch, M_Asset_Class = EQUITIES,
                 M_Asset_Type != ETF, M_Div_Coupon_Class in eligible set.
    Sprint 4 Wave 3: M_Asset_Type gate added. Payout/DivGrowth read from _Pct columns.
    Sprint 4 Wave 6: Momentum (5th input) added.
    """
    cls   = row[cm["M_Div_Coupon_Class"]] if "M_Div_Coupon_Class" in cm else None
    elim  = row[cm["M_Eliminated"]]       if "M_Eliminated"       in cm else None
    asset = row[cm["M_Asset_Class"]]      if "M_Asset_Class"      in cm else None
    atype = row[cm["M_Asset_Type"]]       if "M_Asset_Type"       in cm else None

    blank_scores = {
        "D_QScore_DivSafety":     "",
        "D_QScore_Debt":          "",
        "D_QScore_Profitability": "",
        "D_QScore_Stability":     "",
    }

    insufficient = dict(blank_scores, D_QScore="", D_QTier="Insufficient Data")
    ineligible   = dict(blank_scores, D_QScore="", D_QTier="")

    # Eligibility gates
    if elim == "No Touch":
        return ineligible
    if asset != "EQUITIES":
        return ineligible
    if str(atype or "").strip() == "ETF":  # Wave 3: explicit ETF exclusion
        return ineligible
    if cls not in QSCORE_V2_ELIGIBLE_CLASSES:
        return ineligible

    # Input 1: Streak Proxy (from M_Div_Coupon_Class -- always populated if eligible)
    sc_streak = STREAK_PROXY_V2.get(cls)

    # Input 2: DivGrowth_5Y_Pct (Wave 4 rename)
    dg_raw = row[cm["S_DivGrowth_5Y_Pct"]] if "S_DivGrowth_5Y_Pct" in cm else None
    sc_divgrowth = _score_divgrowth_v2(float(dg_raw)) if dg_raw is not None else None

    # Input 3: PayoutRatio_Pct (Wave 4 rename + x100 scale)
    pr_raw = row[cm["S_PayoutRatio_Pct"]] if "S_PayoutRatio_Pct" in cm else None
    sc_payout = _score_payout_v2(float(pr_raw)) if pr_raw is not None else None

    # Input 4: Beta
    beta_raw = row[cm["S_Beta"]] if "S_Beta" in cm else None
    sc_beta  = _score_beta_v2(float(beta_raw)) if beta_raw is not None else None

    # Input 5: Momentum (Wave 6 -- optional; blank doesn't cause Insufficient Data)
    p5y_raw   = row[cm["S_Price_5Y_Return_Pct"]] if "S_Price_5Y_Return_Pct" in cm else None
    pct52_raw = row[cm["S_BC_Pct52W"]]           if "S_BC_Pct52W"           in cm else None
    sc_momentum = _score_momentum_v2(p5y_raw, pct52_raw)  # None if either blank

    # Any blank core input (1-4) -> Insufficient Data
    core_inputs = [sc_streak, sc_divgrowth, sc_payout, sc_beta]
    if any(s is None for s in core_inputs):
        return insufficient

    # Average: 4 core inputs + Momentum if available
    if sc_momentum is not None:
        all_inputs = core_inputs + [sc_momentum]
    else:
        all_inputs = core_inputs

    score = round(sum(all_inputs) / len(all_inputs), 1)

    if score >= 80:   tier = "A"
    elif score >= 75: tier = "B"
    elif score >= 60: tier = "C"
    else:             tier = "Excluded"

    return dict(blank_scores, D_QScore=score, D_QTier=tier)


# ---------------------------------------------------------------------------
# DIVIDENDS aggregates
# ---------------------------------------------------------------------------

def build_dividend_aggregates(wb):
    """
    Read DIVIDENDS tab, return per-ticker dicts:
      all_time[ticker] = sum of ALL Amount rows, no date filter (inception)
      ltm[ticker]      = rolling 12 months from today (D_AnnualIncome)
      ytd[ticker]      = current calendar year
    """
    ws = wb["DIVIDENDS"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    date_idx   = headers.index("Date")
    amount_idx = headers.index("Amount")
    ticker_idx = headers.index("Ticker")

    today    = date.today()
    ltm_from = today - relativedelta(months=12)
    ytd_from = date(today.year, 1, 1)

    all_time = {}
    ltm      = {}
    ytd      = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[ticker_idx]
        amount = row[amount_idx]
        raw_dt = row[date_idx]

        if not ticker or amount is None:
            continue

        try:
            if isinstance(raw_dt, (datetime, date)):
                dt = raw_dt.date() if isinstance(raw_dt, datetime) else raw_dt
            else:
                dt = datetime.strptime(str(raw_dt), "%Y-%m-%d").date()
        except Exception:
            continue

        amount = float(amount)
        all_time[ticker] = all_time.get(ticker, 0) + amount

        if dt >= ltm_from:
            ltm[ticker] = ltm.get(ticker, 0) + amount

        if dt >= ytd_from:
            ytd[ticker] = ytd.get(ticker, 0) + amount

    return all_time, ltm, ytd


# ---------------------------------------------------------------------------
# Sell Board helpers
# ---------------------------------------------------------------------------

def compute_sell_board(row, cm, total_portfolio_value):
    """Compute all five director signals + D_Sleeve_Weight for one row."""
    cls       = row[cm["M_Div_Coupon_Class"]] if "M_Div_Coupon_Class" in cm else None
    chart     = row[cm["M_Chart_Structure_WTF"]] if "M_Chart_Structure_WTF" in cm else None
    better    = row[cm["M_Better_Candidate"]] if "M_Better_Candidate" in cm else None
    mkt_val   = row[cm["S_MarketValue_GBP"]] if "S_MarketValue_GBP" in cm else None
    # Wave 4: read DivGrowth from renamed column
    div_growth = row[cm["S_DivGrowth_5Y_Pct"]] if "S_DivGrowth_5Y_Pct" in cm else None

    sleeve_weight = (
        float(mkt_val) / total_portfolio_value * 100
        if (mkt_val and total_portfolio_value) else None
    )

    curator = (
        "HOLDS"  if cls in SLEEVE_STREAK_TIERS else
        "REVIEW" if cls in ELIGIBLE_CLASSES    else
        "N/A"
    )

    grower = (
        "FLAG" if (div_growth is not None and float(div_growth) < 0) else
        "PASS" if div_growth is not None else
        "N/A"
    )

    chartist = (
        "CONFIRM" if chart == "DOWNTREND" else
        "NEUTRAL" if chart == "RANGING"   else
        "BLOCK"   if chart == "UPTREND"   else
        "N/A"
    )

    treasurer = (
        "FLAG" if (sleeve_weight is not None and sleeve_weight >= 4.75 and better)
        else "PASS"
    )

    return {
        "D_Curator_Veto":     curator,
        "D_Grower_Signal":    grower,
        "D_Chartist_Signal":  chartist,
        "D_Treasurer_Signal": treasurer,
        "D_Sleeve_Weight":    round(sleeve_weight, 4) if sleeve_weight is not None else "",
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    _ctx["run_id"] = str(uuid.uuid4())[:8]
    _ctx["uid"]    = ""

    start = datetime.now()
    _log("info", "STARTUP", "RUN_START", "", "derive_engine.py starting")

    wb_path = find_workbook()
    wb      = openpyxl.load_workbook(wb_path, keep_vba=True, data_only=True)
    ws_uni  = wb["Universe"]

    headers = [c.value for c in next(ws_uni.iter_rows(min_row=1, max_row=1))]
    cm = {h: i for i, h in enumerate(headers) if h}

    # -- Step 1: DIVIDENDS aggregates -----------------------------------------
    _log("info", "DERIVE", "DIVIDENDS_AGG", "", "Building dividend aggregates")
    all_time_map, ltm_map, ytd_map = build_dividend_aggregates(wb)

    all_rows = list(ws_uni.iter_rows(min_row=2, values_only=True))
    _log("info", "DERIVE", "UNIVERSE_LOAD", "", f"{len(all_rows)} Universe rows loaded")

    computed = [{} for _ in all_rows]

    total_portfolio_value = 0.0
    annual_income_by_row  = {}

    for i, row in enumerate(all_rows):
        ticker = row[cm["M_Ticker"]] if "M_Ticker" in cm else None
        if not ticker:
            continue

        all_income = all_time_map.get(ticker, 0)
        ann_income = ltm_map.get(ticker, 0)
        ytd_income = ytd_map.get(ticker, 0)

        computed[i]["D_All_Income"]     = all_income if all_income else ""
        computed[i]["D_AnnualIncome"]   = ann_income if ann_income else ""
        computed[i]["D_Income_LTM_GBP"] = ann_income if ann_income else ""
        computed[i]["D_Income_YTD_GBP"] = ytd_income if ytd_income else ""
        annual_income_by_row[i] = ann_income

        # D_Income_Payback_Pct = D_All_Income / S_CostBasis x 100
        # Wave 3: stored as x100 (raw %) per _Pct convention.
        cost_basis = row[cm["S_CostBasis"]] if "S_CostBasis" in cm else None
        if all_income and cost_basis:
            try:
                computed[i]["D_Income_Payback_Pct"] = round(
                    all_income / float(cost_basis) * 100, 4
                )
            except (TypeError, ValueError, ZeroDivisionError):
                computed[i]["D_Income_Payback_Pct"] = ""
        else:
            computed[i]["D_Income_Payback_Pct"] = ""

        mkt_val = row[cm["S_MarketValue_GBP"]] if "S_MarketValue_GBP" in cm else None
        if mkt_val:
            total_portfolio_value += float(mkt_val)

    _log("info", "DERIVE", "PORTFOLIO_VALUE", "",
         f"Total portfolio value: {total_portfolio_value:,.2f}")

    # -- Step 1b: Read existing S_QScore for shift chain ----------------------
    # Sprint 4 Wave 3: proper 3-column shift chain.
    # Read existing S_QScore from the workbook (last confirmed written value).
    # After QScore is computed: if D_QScore_new != S_QScore_existing:
    #   S_QScore_Prior = S_QScore_existing; S_QScore = D_QScore_new
    existing_s_qscore = {}
    if "S_QScore" in cm:
        for i, row in enumerate(all_rows):
            existing_s_qscore[i] = row[cm["S_QScore"]]  # may be None/""/numeric

    # -- Step 2: QScore --------------------------------------------------------
    _log("info", "DERIVE", "QSCORE", "", "Computing QScore v2")
    scored = ineligible_count = insufficient_count = 0
    for i, row in enumerate(all_rows):
        result = compute_qscore(row, cm)
        computed[i].update(result)
        if result.get("D_QScore") not in ("", None) and result.get("D_QTier") not in ("", "Insufficient Data"):
            scored += 1
        elif result.get("D_QTier") == "Insufficient Data":
            insufficient_count += 1
        else:
            ineligible_count += 1

    _log("info", "DERIVE", "QSCORE", "",
         f"{scored} scored, {insufficient_count} insufficient, {ineligible_count} ineligible")

    # -- Step 2b: QScore 3-column shift chain (Wave 3) -------------------------
    # D_QScore = computed live (always written).
    # S_QScore = last confirmed written value.
    # S_QScore_Prior = previous S_QScore before last change.
    # Shift only when D_QScore differs from S_QScore.
    for i in range(len(all_rows)):
        new_d_qscore = computed[i].get("D_QScore")
        ex_s_qscore  = existing_s_qscore.get(i)

        # Normalise: "" and None both treated as no value
        new_val = None if (new_d_qscore == "" or new_d_qscore is None) else new_d_qscore
        ex_val  = None if (ex_s_qscore  == "" or ex_s_qscore  is None) else ex_s_qscore

        if new_val is not None and new_val != ex_val:
            # Shift: S_QScore_Prior = old S_QScore; S_QScore = new D_QScore
            if ex_val is not None:
                computed[i]["S_QScore_Prior"] = ex_val
            computed[i]["S_QScore"] = new_val
        # else: unchanged -- conditional write guard handles no-op

    # -- Step 3: Price derived -------------------------------------------------
    _log("info", "DERIVE", "PRICE_DERIVED", "", "Computing price-derived columns")
    for i, row in enumerate(all_rows):
        elim = row[cm["M_Eliminated"]] if "M_Eliminated" in cm else None
        if elim == "No Touch":
            computed[i].update({
                "D_PE_Discount":       "",
                "D_Yield_Premium_Pct": "",
                "D_Price_vs_52W_High": "",
            })
            continue

        pe     = row[cm["S_PE_Ratio"]]         if "S_PE_Ratio"         in cm else None
        pe_avg = row[cm["S_PE_5YAvg"]]          if "S_PE_5YAvg"         in cm else None
        # Wave 4: renamed columns (stored x100, e.g. 4.0 = 4%)
        yld    = row[cm["S_Dividend_Yield_Pct"]] if "S_Dividend_Yield_Pct" in cm else None
        yavg   = row[cm["S_DivYield_5YAvg_Pct"]] if "S_DivYield_5YAvg_Pct" in cm else None
        price  = row[cm["S_Last_Price"]]         if "S_Last_Price"         in cm else None
        high   = row[cm["S_52W_High"]]           if "S_52W_High"           in cm else None

        pe_disc  = (
            round((float(pe) - float(pe_avg)) / float(pe_avg), 4)
            if (pe and pe_avg and float(pe_avg) != 0) else ""
        )
        # D_Yield_Premium_Pct: both values stored x100, so difference is also x100 points
        yld_prem = (
            round(float(yld) - float(yavg), 4)
            if (yld is not None and yavg is not None) else ""
        )
        p52 = (
            round((float(price) - float(high)) / float(high), 4)
            if (price and high and float(high) != 0) else ""
        )

        computed[i].update({
            "D_PE_Discount":       pe_disc,
            "D_Yield_Premium_Pct": yld_prem,
            "D_Price_vs_52W_High": p52,
        })

    # -- Step 4: Sell Board signals --------------------------------------------
    _log("info", "DERIVE", "SELL_BOARD", "", "Computing Sell Board signals")
    for i, row in enumerate(all_rows):
        elim = row[cm["M_Eliminated"]] if "M_Eliminated" in cm else None
        if elim == "No Touch":
            computed[i].update({k: "" for k in [
                "D_Curator_Veto", "D_Grower_Signal", "D_Chartist_Signal",
                "D_Treasurer_Signal", "D_Sleeve_Weight",
            ]})
            continue
        sb = compute_sell_board(row, cm, total_portfolio_value)
        computed[i].update(sb)

    # -- Step 5: Income / YOC + D_GBP_Per_GBP1_Div ----------------------------
    _log("info", "DERIVE", "INCOME_YOC", "", "Computing income/YOC columns")
    gbp_per_gbp1 = {}
    for i, row in enumerate(all_rows):
        elim = row[cm["M_Eliminated"]] if "M_Eliminated" in cm else None
        if elim == "No Touch":
            computed[i].update({k: "" for k in [
                "D_YOC", "D_YoC_PerHolding", "D_GBP_Per_GBP1_Div",
            ]})
            continue

        cost_bas = row[cm["S_CostBasis"]]          if "S_CostBasis"          in cm else None
        # Wave 4: S_Dividend_Yield renamed to S_Dividend_Yield_Pct (stored x100)
        yld      = row[cm["S_Dividend_Yield_Pct"]] if "S_Dividend_Yield_Pct" in cm else None
        ann_inc  = annual_income_by_row.get(i, 0)

        yoc         = round(float(yld), 4) if yld is not None else ""
        yoc_holding = (
            round(ann_inc / float(cost_bas), 4)
            if (ann_inc and cost_bas) else ""
        )
        gbp_eff = (
            round(float(cost_bas) / ann_inc, 2)
            if (ann_inc and cost_bas and ann_inc > 0) else ""
        )

        computed[i].update({
            "D_YOC":              yoc,
            "D_YoC_PerHolding":   yoc_holding,
            "D_GBP_Per_GBP1_Div": gbp_eff,
        })
        if isinstance(gbp_eff, float):
            gbp_per_gbp1[i] = gbp_eff

    # -- Comptroller rank (across DIV_CORE only) --------------------------------
    _log("info", "DERIVE", "COMPTROLLER", "", "Computing Comptroller rank")
    div_core_vals = {
        i: v for i, v in gbp_per_gbp1.items()
        if "M_Sleeve" in cm and all_rows[i][cm["M_Sleeve"]] == "DIV_CORE"
    }
    if div_core_vals:
        sorted_vals   = sorted(div_core_vals.values())
        n             = len(sorted_vals)
        p75_threshold = sorted_vals[int(n * 0.75)]

        for i, row in enumerate(all_rows):
            sleeve = row[cm["M_Sleeve"]] if "M_Sleeve" in cm else None
            if sleeve != "DIV_CORE" or i not in gbp_per_gbp1:
                computed[i].update({
                    "D_Comptroller_Rank": "", "D_Comptroller_Signal": "",
                })
                continue
            val    = gbp_per_gbp1[i]
            rank   = sorted_vals.index(val) + 1
            signal = "FLAG" if val >= p75_threshold else "PASS"
            computed[i].update({
                "D_Comptroller_Rank":   rank,
                "D_Comptroller_Signal": signal,
            })
    else:
        for i in range(len(all_rows)):
            computed[i].update({
                "D_Comptroller_Rank": "", "D_Comptroller_Signal": "",
            })

    # -- Board verdict ---------------------------------------------------------
    flag_cols = [
        "D_Comptroller_Signal", "D_Grower_Signal",
        "D_Chartist_Signal", "D_Treasurer_Signal",
    ]
    for i in range(len(all_rows)):
        curator = computed[i].get("D_Curator_Veto", "")
        if not curator:
            computed[i].update({"D_Board_Flags": "", "D_Board_Verdict": "N/A"})
            continue
        flags = sum(1 for k in flag_cols if computed[i].get(k) == "FLAG")
        computed[i]["D_Board_Flags"]   = flags
        computed[i]["D_Board_Verdict"] = (
            "VETOED"      if curator == "HOLDS" else
            "SELL REVIEW" if flags >= 3          else
            "HOLD"
        )

    # -- Step 6: Bolt-ons ------------------------------------------------------
    for i in range(len(all_rows)):
        computed[i]["D_ValueClass"] = ""  # TBD -- definition pending Kyle

    # -- Write all computed values back to sheet --------------------------------
    _log("info", "DERIVE", "WRITE_START", "", "Writing D_ values to Universe sheet")
    written = skipped = 0

    for i, vals in enumerate(computed):
        excel_row = i + 2  # row 1 = header
        ticker    = all_rows[i][cm["M_Ticker"]] if "M_Ticker" in cm else ""
        for col_name, value in vals.items():
            if col_name not in cm:
                skipped += 1
                continue
            col_idx = cm[col_name] + 1  # 0-indexed cm -> 1-indexed openpyxl
            write_cell(ws_uni, excel_row, col_idx, value)
            written += 1
        if ticker and written % 50 == 0:
            _log("info", "DERIVE", "WRITE_PROGRESS", str(ticker),
                 f"row {excel_row} written")

    _log("info", "DERIVE", "WRITE_END", "",
         f"{written} cells written, {skipped} col(s) not in sheet (skipped)")

    # -- Save ------------------------------------------------------------------
    save_workbook_with_increment(wb, wb_path)
    _log("info", "COMPLETE", "RUN_END", "", "Workbook saved")

    end     = datetime.now()
    elapsed = str(end - start).split(".")[0]
    _log("info", "COMPLETE", "TIMING", "",
         f"Started {start.strftime('%Y-%m-%d %H:%M:%S')} | "
         f"Ended {end.strftime('%H:%M:%S')} | Duration {elapsed}")


if __name__ == "__main__":
    run()
