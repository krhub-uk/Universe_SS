"""
sync_engine_tv.py — TradingView watchlist .txt generator
Sprint 3 | Spec: V4.6 §11

Two independent watchlist-generation mechanisms:

1. Free-text routing (M_Export_TV):
   - Value = literal destination watchlist name
   - "N" = exclude
   - "Y" WARNING: treated as watchlist named "Y" — populate with
     actual watchlist name if different behaviour is intended.

2. Boolean-flag auto-derivation:
   - M_Sleeve_Watchlist = "Y"          → one .txt per distinct M_Sleeve value
   - M_Universe_Watchlist = "Y"        → one .txt per distinct M_Universe value
   - M_Div_Coupon_Class_Watchlist = "Y"→ one .txt per distinct M_Div_Coupon_Class value
   (Columns must exist in sheet; absent columns are skipped with a WARNING.)

Section headers within each file:
   ####<M_Related_To value> — distinct values of M_Related_To for included tickers.
   Tickers with no M_Related_To value go into an "####Other" section at the end.

Output: Cowork/Output/TV/<watchlist_name>.txt — overwrite on every run.

Standing rules:
   - M_Eliminated = "No Touch" → excluded from all watchlists, checked first
   - S_TV_Ticker is the send value. If blank, M_Ticker is used as fallback.
   - .txt files are fully disposable. Spreadsheet = sole source of truth.
   - No cadence logic — generates all configured watchlists every run.
   - Human chooses which files to upload to TV.

Spec: GDrive/Claude/TradingUniverse/00_Portfolio_Automation_Spec_V4.6.md §11
"""

import os
import sys
import logging
import re
from collections import defaultdict

import openpyxl
from workbook_io import find_workbook

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "Output", "TV")
GDRIVE_DEST = "gdrive:Claude/TradingUniverse/Watchlists/TV"
LOG_FILE   = "/var/log/portfolio/sync_engine_tv.log"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# Boolean-flag columns → source column mapping
FLAG_COLUMNS = {
    "M_Sleeve_Watchlist":           "M_Sleeve",
    "M_Universe_Watchlist":         "M_Universe",
    "M_Div_Coupon_Class_Watchlist": "M_Div_Coupon_Class",
}


def safe_filename(name):
    """Convert watchlist name to a safe filename."""
    name = str(name).strip()
    name = re.sub(r'[^\w\s\-]', '', name)
    name = re.sub(r'\s+', '_', name)
    return name


def build_watchlists(ws, headers):
    """
    Returns dict: watchlist_name → list of (section_label, ticker)
    Preserves insertion order within each watchlist.
    """
    cm = {h: i for i, h in enumerate(headers) if h}

    def get(row, col):
        idx = cm.get(col)
        return row[idx] if idx is not None and idx < len(row) else None

    # Check which flag columns actually exist
    available_flags = {
        flag: source for flag, source in FLAG_COLUMNS.items()
        if flag in cm
    }
    missing_flags = set(FLAG_COLUMNS) - set(available_flags)
    if missing_flags:
        log.warning(f"Boolean-flag columns not found in sheet (skipping): {missing_flags}")

    # watchlist_name → {section_label → [ticker, ...]}
    watchlists = defaultdict(lambda: defaultdict(list))

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker_raw = get(row, 'M_Ticker')
        if not ticker_raw:
            continue

        # M_Eliminated gate — first check
        if get(row, 'M_Eliminated') == 'No Touch':
            continue

        # TV ticker — S_TV_Ticker preferred, M_Ticker fallback
        tv_ticker = get(row, 'S_TV_Ticker') or ticker_raw
        section   = get(row, 'M_Related_To') or 'Other'

        # ── Mechanism 1: free-text routing ────────────────────────────────
        # N = exclude from free-text only. Y = no named watchlist (boolean flags handle it).
        # Any other string = literal watchlist name.
        # M_Export_TV does NOT gate Mechanism 2 — the two mechanisms are independent.
        export_tv = get(row, 'M_Export_TV')
        export_tv_str = str(export_tv).strip().upper() if export_tv else ''
        if export_tv_str not in ('', 'N', 'Y'):
            watchlists[str(export_tv).strip()][section].append(tv_ticker)

        # ── Mechanism 2: boolean-flag auto-derivation ─────────────────────
        # Runs for ALL non-eliminated tickers — independent of M_Export_TV.
        # Each flag column independently controls inclusion in its derived watchlist.
        for flag_col, source_col in available_flags.items():
            flag_val = get(row, flag_col)
            if str(flag_val).strip().upper() == 'Y':
                source_val = get(row, source_col)
                if source_val:
                    watchlists[str(source_val).strip()][section].append(tv_ticker)

    return watchlists


def write_watchlist_file(name, sections, output_dir):
    """Write one .txt file for a watchlist, grouped by section. Deduplicates tickers globally."""
    filename = safe_filename(name) + ".txt"
    filepath = os.path.join(output_dir, filename)

    lines = []
    seen = set()  # deduplicate across AND within all sections in this file
    ordered = sorted(sections.keys(), key=lambda s: (s == 'Other', s))
    for section in ordered:
        section_tickers = []
        for t in sections[section]:
            if t not in seen:
                section_tickers.append(t)
                seen.add(t)  # update immediately to catch within-section dupes
        if not section_tickers:
            continue
        lines.append(f"####{section}")
        lines.extend(section_tickers)

    with open(filepath, 'w') as f:
        f.write('\n'.join(lines) + '\n')

    return filepath, len(seen)


def run():
    from datetime import datetime
    start = datetime.now()
    log.info("sync_engine_tv.py started")

    wb_path = find_workbook()
    log.info(f"Workbook: {wb_path}")

    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True, keep_vba=True)
    ws = wb['Universe']
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    log.info("Building watchlists...")
    watchlists = build_watchlists(ws, headers)

    if not watchlists:
        log.warning("No watchlists generated — check M_Export_TV values and flag columns.")
    else:
        log.info(f"Watchlists to generate: {sorted(watchlists.keys())}")

    # Clear output folder before writing — removes stale files from previous runs
    # (prevents case variants like Core.txt + CORE.txt coexisting)
    import glob
    stale = glob.glob(os.path.join(OUTPUT_DIR, "*.txt"))
    for f in stale:
        os.remove(f)
    if stale:
        log.info(f"Cleared {len(stale)} stale file(s) from output folder")

    total_files = 0
    total_tickers = 0

    for name, sections in sorted(watchlists.items()):
        ticker_count = sum(len(t) for t in sections.values())
        filepath, line_count = write_watchlist_file(name, sections, OUTPUT_DIR)
        log.info(f"  Written: {os.path.basename(filepath)} ({ticker_count} tickers)")
        total_files   += 1
        total_tickers += ticker_count

    log.info(f"Done: {total_files} files, {total_tickers} ticker entries total")
    log.info(f"Output folder: {OUTPUT_DIR}")

    # Push to GDrive
    import subprocess
    log.info(f"Pushing to GDrive: {GDRIVE_DEST}")
    result = subprocess.run(
        ["rclone", "copy", OUTPUT_DIR, GDRIVE_DEST, "--include", "*.txt"],
        capture_output=True, text=True
    )
    if result.returncode == 0:
        log.info("GDrive push complete")
    else:
        log.warning(f"GDrive push failed: {result.stderr.strip()}")

    end = datetime.now()
    elapsed = str(end - start).split(".")[0]
    log.info(
        f"--- Started: {start.strftime('%Y-%m-%d %H:%M:%S')} | "
        f"Ended: {end.strftime('%H:%M:%S')} | Duration: {elapsed} ---"
    )


if __name__ == "__main__":
    run()
