"""
coverage_script.py -- SECTOR_COVERAGE + ALLOC_COVERAGE tab population
Spec: 00_Portfolio_Automation_Spec_V5.5.md, Wave 7. Sprint 4.

Runs: weekly cadence (called from run_weekly.sh after derive_engine.py).

Writes to two separate tabs (user-created in Excel):
  SECTOR_COVERAGE  -- one row per sector, S_Invested="Y" rows only
  ALLOC_COVERAGE   -- one row per sleeve, S_Invested="Y" rows only

Both tabs: value-only clear on each run (preserves user formatting,
colours, borders). Headers written to row 1; data from row 2 onwards.
Named ranges defined for chart anchoring after each write.

SECTOR_COVERAGE columns (DB-friendly, row 1 headers):
  sector | holdings_count | mkt_value_gbp | portfolio_pct |
  cost_gbp | gl_gbp | gl_pct

ALLOC_COVERAGE columns (DB-friendly, row 1 headers):
  sleeve | holdings_count | mkt_value_gbp | portfolio_pct |
  target_pct | vs_target_pct | cost_gbp | gl_gbp

S_Invested = "Y" filter applied throughout.
Sleeve target % from Lookups Allocations table: Allocations / Alloc_Pct columns.
target_pct stored as raw % (e.g. 35.0 = 35%); vs_target_pct = portfolio_pct - target_pct.

Wave 8: Structured PHASE/ACTION/TICKER/RUN_ID/UID logging.
"""

import logging
import sys
import uuid
from datetime import datetime

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

from workbook_io import (
    BASE_PATH,
    find_workbook,
    save_workbook_with_increment,
    write_cell,
)

# ---------------------------------------------------------------------------
# Logging (Wave 8)
# ---------------------------------------------------------------------------

LOG_DIR = BASE_PATH / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "coverage_script.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("coverage_script")

_ctx: dict = {}
_VOCAB_KEYS = ("PHASE=", "ACTION=", "TICKER=", "RUN_ID=", "UID=")


def _log(level: str, phase: str, action: str, ticker: str, message: str) -> None:
    run_id = _ctx.get("run_id", "")
    uid    = _ctx.get("uid", "")
    line   = (
        f"PHASE={phase} ACTION={action} TICKER={ticker} "
        f"RUN_ID={run_id} UID={uid} | {message}"
    )
    for key in _VOCAB_KEYS:
        if key not in line:
            log.warning(f"[VOCAB_FAIL] Missing {key}")
    getattr(log, level.lower())(line)


# ---------------------------------------------------------------------------
# Column definitions (DB-friendly headers)
# ---------------------------------------------------------------------------

# SECTOR_COVERAGE tab
SECTOR_HEADERS = [
    "sector", "holdings_count", "mkt_value_gbp", "portfolio_pct",
    "cost_gbp", "gl_gbp", "gl_pct",
]

# ALLOC_COVERAGE tab
ALLOC_HEADERS = [
    "sleeve", "holdings_count", "mkt_value_gbp", "portfolio_pct",
    "target_pct", "vs_target_pct", "unallocated_gbp", "cost_gbp", "gl_gbp",
]

SECTOR_TAB  = "SECTOR_COVERAGE"
ALLOC_TAB   = "ALLOC_COVERAGE"

SECTOR_NAMED_RANGE = "SECTOR_COVERAGE_DATA"
ALLOC_NAMED_RANGE  = "ALLOC_COVERAGE_DATA"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(v, default=0.0):
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _clear_tab_values(ws):
    """
    Value-only clear: set cell values to None from row 2 onwards.
    Does NOT delete rows — preserves all formatting, colours, borders.
    Clears up to current max_row so stale rows from prior runs are blanked.
    """
    if ws.max_row < 2:
        return
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def _write_headers(ws, headers):
    """Write DB-friendly header row (row 1). Direct value set — no font size change."""
    for col_i, hdr in enumerate(headers, start=1):
        ws.cell(row=1, column=col_i).value = hdr


def _set_named_range(wb, name, tab, first_data_row, last_data_row, num_cols):
    """
    Define (or replace) a named range spanning the data area of a tab.
    Used as a stable anchor for charts regardless of row count changes.
    """
    if last_data_row < first_data_row:
        return  # no data rows — skip
    last_col_letter = openpyxl.utils.get_column_letter(num_cols)
    ref = f"'{tab}'!$A${first_data_row}:${last_col_letter}${last_data_row}"
    defn = DefinedName(name, attr_text=ref)
    # Remove stale definition if it exists
    try:
        if name in wb.defined_names:
            wb.defined_names.delete(name)
    except (AttributeError, KeyError):
        pass
    try:
        wb.defined_names[name] = defn
    except Exception:
        wb.defined_names.append(defn)


# ---------------------------------------------------------------------------
# Load Universe data
# ---------------------------------------------------------------------------

def header_map(ws):
    """{header_name: 0-based_index} from row 1 values."""
    return {
        cell.value: cell.column - 1
        for cell in ws[1]
        if cell.value is not None
    }


def load_universe_rows(wb):
    """
    Returns a list of dicts for every Universe row where S_Invested = "Y".
    """
    ws = wb["Universe"]
    cm = header_map(ws)

    needed = [
        "S_Invested", "S_Sector", "M_Sleeve",
        "S_MarketValue_GBP", "S_CostBasis", "S_PnL_GBP", "S_PnL_Pct",
    ]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        invested = row[cm["S_Invested"]] if "S_Invested" in cm else None
        if invested != "Y":
            continue
        rows.append({
            "S_Sector":          row[cm["S_Sector"]]          if "S_Sector"          in cm else None,
            "M_Sleeve":          row[cm["M_Sleeve"]]          if "M_Sleeve"          in cm else None,
            "S_MarketValue_GBP": row[cm["S_MarketValue_GBP"]] if "S_MarketValue_GBP" in cm else None,
            "S_CostBasis":       row[cm["S_CostBasis"]]       if "S_CostBasis"       in cm else None,
            "S_PnL_GBP":         row[cm["S_PnL_GBP"]]         if "S_PnL_GBP"         in cm else None,
            "S_PnL_Pct":         row[cm["S_PnL_Pct"]]         if "S_PnL_Pct"         in cm else None,
        })

    return rows


# ---------------------------------------------------------------------------
# Load Lookups Allocations table
# ---------------------------------------------------------------------------

def load_allocations(wb):
    """
    Reads Lookups tab Allocations table by column name.
    Returns dict: {sleeve_name: target_pct_raw} where target_pct_raw is 35 (= 35%).
    """
    ws   = wb["Lookups"]
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, row in enumerate(rows):
        if row and "Allocations" in row and "Alloc_Pct" in row:
            header_row_idx = i
            break

    if header_row_idx is None:
        _log("warning", "COVERAGE", "ALLOC_MISSING", "",
             "Lookups tab: could not find Allocations/Alloc_Pct header row")
        return {}

    header    = list(rows[header_row_idx])
    alloc_col = header.index("Allocations")
    pct_col   = header.index("Alloc_Pct")

    allocations = {}
    for row in rows[header_row_idx + 1:]:
        if not row or row[alloc_col] is None:
            break
        sleeve = row[alloc_col]
        pct    = row[pct_col]
        if sleeve and pct is not None:
            allocations[str(sleeve).strip()] = _safe_float(pct)

    return allocations


# ---------------------------------------------------------------------------
# Aggregate helpers
# ---------------------------------------------------------------------------

def _aggregate(filtered_rows):
    """Sum holdings, mkt_val, cost, pnl_gbp for a list of rows."""
    holdings = len(filtered_rows)
    mkt_val  = sum(_safe_float(r["S_MarketValue_GBP"]) for r in filtered_rows)
    cost     = sum(_safe_float(r["S_CostBasis"])        for r in filtered_rows)
    pnl_gbp  = sum(_safe_float(r["S_PnL_GBP"])          for r in filtered_rows)
    return holdings, mkt_val, cost, pnl_gbp


def _pnl_pct(cost, pnl_gbp):
    if cost and cost != 0:
        return round(pnl_gbp / cost * 100, 2)
    return None


# ---------------------------------------------------------------------------
# SECTOR COVERAGE build
# ---------------------------------------------------------------------------

SECTOR_ORDER = [
    "Basic Materials", "Communication Services", "Consumer Cyclical",
    "Consumer Defensive", "Energy", "Financial Services",
    "Healthcare", "Industrials", "Real Estate",
    "Technology", "Utilities",
    "Unknown", "Other",
]


def build_sector_coverage(all_rows):
    """
    Returns list of dicts, one per sector, plus TOTAL row.
    Columns: sector, holdings_count, mkt_value_gbp, portfolio_pct,
             cost_gbp, gl_gbp, gl_pct
    """
    sector_map = {}
    for row in all_rows:
        sector = str(row["S_Sector"] or "Unknown").strip() or "Unknown"
        sector_map.setdefault(sector, []).append(row)

    known  = [s for s in SECTOR_ORDER if s in sector_map]
    extra  = sorted(s for s in sector_map if s not in SECTOR_ORDER)
    order  = known + extra

    total_mkt = sum(_safe_float(r["S_MarketValue_GBP"]) for r in all_rows)

    result = []
    for sector in order:
        rows = sector_map[sector]
        holdings, mkt_val, cost, pnl_gbp = _aggregate(rows)
        portfolio_pct = round(mkt_val / total_mkt * 100, 2) if total_mkt else None
        result.append({
            "sector":        sector,
            "holdings_count": holdings,
            "mkt_value_gbp": round(mkt_val, 2),
            "portfolio_pct": portfolio_pct,
            "cost_gbp":      round(cost, 2),
            "gl_gbp":        round(pnl_gbp, 2),
            "gl_pct":        _pnl_pct(cost, pnl_gbp),
        })

    # TOTAL row
    t_holdings, t_mkt, t_cost, t_pnl = _aggregate(all_rows)
    result.append({
        "sector":        "TOTAL",
        "holdings_count": t_holdings,
        "mkt_value_gbp": round(t_mkt, 2),
        "portfolio_pct": 100.0 if all_rows else None,
        "cost_gbp":      round(t_cost, 2),
        "gl_gbp":        round(t_pnl, 2),
        "gl_pct":        _pnl_pct(t_cost, t_pnl),
    })

    return result


# ---------------------------------------------------------------------------
# ALLOC COVERAGE build
# ---------------------------------------------------------------------------

def build_alloc_coverage(all_rows, allocations):
    """
    Returns list of dicts, one per sleeve, plus TOTAL row.
    Columns: sleeve, holdings_count, mkt_value_gbp, portfolio_pct,
             target_pct, vs_target_pct, cost_gbp, gl_gbp

    Sleeve order driven by Lookups Allocations; unknown sleeves appended
    alphabetically (e.g. DIAMOND will appear if M_Sleeve = "DIAMOND").
    target_pct = raw % from Alloc_Pct (e.g. 35.0 = 35%).
    vs_target_pct = portfolio_pct - target_pct.
    """
    total_mkt  = sum(_safe_float(r["S_MarketValue_GBP"]) for r in all_rows)

    sleeve_map = {}
    for row in all_rows:
        sleeve = str(row["M_Sleeve"] or "Unknown").strip() or "Unknown"
        sleeve_map.setdefault(sleeve, []).append(row)

    known_sleeves  = list(allocations.keys())
    extra_sleeves  = sorted(s for s in sleeve_map if s not in allocations)
    ordered        = known_sleeves + extra_sleeves

    result = []
    for sleeve in ordered:
        rows = sleeve_map.get(sleeve, [])
        holdings, mkt_val, cost, pnl_gbp = _aggregate(rows)
        portfolio_pct = round(mkt_val / total_mkt * 100, 2) if total_mkt else None
        target_raw    = allocations.get(sleeve)           # e.g. 35.0 (= 35%)
        vs_target_pct  = (
            round(portfolio_pct - target_raw, 2)
            if (portfolio_pct is not None and target_raw is not None)
            else None
        )
        unallocated_gbp = (
            round((target_raw / 100 * total_mkt) - mkt_val, 2)
            if target_raw is not None
            else None
        )
        result.append({
            "sleeve":          sleeve,
            "holdings_count":  holdings,
            "mkt_value_gbp":   round(mkt_val, 2),
            "portfolio_pct":   portfolio_pct,
            "target_pct":      target_raw,
            "vs_target_pct":   vs_target_pct,
            "unallocated_gbp": unallocated_gbp,
            "cost_gbp":        round(cost, 2),
            "gl_gbp":          round(pnl_gbp, 2),
        })

    # TOTAL row
    t_holdings, t_mkt, t_cost, t_pnl = _aggregate(all_rows)
    result.append({
        "sleeve":          "TOTAL",
        "holdings_count":  t_holdings,
        "mkt_value_gbp":   round(t_mkt, 2),
        "portfolio_pct":   100.0 if all_rows else None,
        "target_pct":      None,
        "vs_target_pct":   None,
        "unallocated_gbp": None,
        "cost_gbp":        round(t_cost, 2),
        "gl_gbp":          round(t_pnl, 2),
    })

    return result


# ---------------------------------------------------------------------------
# Write helpers
# ---------------------------------------------------------------------------

def _write_tab(wb, tab_name, headers, data_rows, named_range):
    """
    Generic write-to-tab:
      1. Value-only clear (preserve formatting)
      2. Write DB-friendly headers to row 1 (value only, no font change)
      3. Write data rows 2..N using write_cell (conditional, preserves alignment)
      4. Define named range over data area (rows 2..N)
    """
    if tab_name not in wb.sheetnames:
        _log("warning", "COVERAGE", "TAB_MISSING", "",
             f"Tab '{tab_name}' not found — skipping. Create it in Excel first.")
        return

    ws = wb[tab_name]

    # Step 1: value-only clear
    _clear_tab_values(ws)

    # Step 2: headers (direct value assignment — don't alter header row formatting)
    _write_headers(ws, headers)

    # Step 3: data rows
    data_row_start = 2
    current_row    = data_row_start
    for row_data in data_rows:
        for col_i, key in enumerate(headers, start=1):
            write_cell(ws, current_row, col_i, row_data.get(key))
        current_row += 1

    last_data_row = current_row - 1

    # Step 4: named range
    _set_named_range(wb, named_range, tab_name,
                     data_row_start, last_data_row, len(headers))
    _log("info", "COVERAGE", "NAMED_RANGE", "",
         f"{named_range} => '{tab_name}'!$A${data_row_start}:${openpyxl.utils.get_column_letter(len(headers))}${last_data_row}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run():
    _ctx["run_id"] = str(uuid.uuid4())[:8]
    _ctx["uid"]    = ""

    start = datetime.now()
    _log("info", "STARTUP", "RUN_START", "", "coverage_script.py starting")

    wb_path = find_workbook()
    wb      = openpyxl.load_workbook(wb_path, keep_vba=True)

    _log("info", "COVERAGE", "UNIVERSE_LOAD", "", "Loading Universe invested rows")
    all_rows = load_universe_rows(wb)
    _log("info", "COVERAGE", "UNIVERSE_LOAD", "", f"{len(all_rows)} S_Invested=Y rows loaded")

    _log("info", "COVERAGE", "ALLOC_LOAD", "", "Loading Lookups Allocations table")
    allocations = load_allocations(wb)
    _log("info", "COVERAGE", "ALLOC_LOAD", "", f"{len(allocations)} sleeve allocations loaded")

    # Build data
    _log("info", "COVERAGE", "SECTOR_BUILD", "", "Building SECTOR_COVERAGE data")
    sector_rows = build_sector_coverage(all_rows)

    _log("info", "COVERAGE", "ALLOC_BUILD", "", "Building ALLOC_COVERAGE data")
    alloc_rows = build_alloc_coverage(all_rows, allocations)

    # Write tabs
    _log("info", "COVERAGE", "WRITE_START", "", f"Writing {SECTOR_TAB}")
    _write_tab(wb, SECTOR_TAB, SECTOR_HEADERS, sector_rows, SECTOR_NAMED_RANGE)

    _log("info", "COVERAGE", "WRITE_START", "", f"Writing {ALLOC_TAB}")
    _write_tab(wb, ALLOC_TAB, ALLOC_HEADERS, alloc_rows, ALLOC_NAMED_RANGE)

    save_workbook_with_increment(wb, wb_path)

    elapsed = str(datetime.now() - start).split(".")[0]
    _log("info", "COMPLETE", "RUN_END", "",
         f"{len(sector_rows)} sector rows, {len(alloc_rows)} alloc rows. Duration {elapsed}")


if __name__ == "__main__":
    run()
