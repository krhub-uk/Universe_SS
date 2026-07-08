"""
Price Action Script — EOD cadence
Spec: 00_Portfolio_Automation_Spec_V3.8.md, Section 7b. Sprint 2 script
re-engineering: replaces the EOD scope of price_action_4h.py.

Runs: EOD (11pm).

Scope each run:
    M_Eliminated != "No Touch"   (universal kill switch, checked first)
    -- all rows, all sleeves. No M_Fetch_Cadence filter (same posture as
    price_action_intraday.py -- §7b).

S_ columns fed -- all 9 intraday attributes (see price_action_intraday.py)
PLUS:
    S_Daily_Close, S_LastClose_Volume, S_1W_Change_%, S_2W_Change_%,
    S_52W_High, S_52W_Low

D_ columns computed this run (after price fetch):
    D_Volume_Delta_%, D_Volume_Flag, D_Cap_Candle_Mid, D_Close_Direction,
    D_Context_Flag, D_Cap_Mid_Anchor, D_Cap_Anchor_Date,
    D_Cap_Anchor_Active, D_Price_vs_Cap_Mid, D_Price_vs_52W_High

Source: one yfinance Ticker object per row --
    history(period="1mo") for change-% calcs (see note below),
    history(period="1d") for OHLCV,
    .info[] ONLY for S_52W_High / S_52W_Low (fiftyTwoWeekHigh/fiftyTwoWeekLow).

After the price fetch completes for all in-scope rows: check Inputs/HL and
Inputs/BC for any present CSV files. If any are found, call
portfolio_csv_ingestion.run() directly (imported, not subprocessed). If none
are present, skip silently -- no error, no log noise beyond a one-line note.

Notes on scope decisions not fully specified in §7b itself:
  - §7b's literal text says history(period="5d") for the change-% calcs.
    A 5-day pull can support S_1D/2D/3D_Change_% (same as
    price_action_intraday.py) and marginally S_1W_Change_% (needs 6 rows;
    "5d" typically returns ~5), but cannot support S_2W_Change_% at all --
    a 2-week-back close needs ~10 trading days of history, which a 5-day
    pull structurally cannot contain, regardless of how the comparison is
    indexed. Raised with the user directly; resolved as: pull
    history(period="1mo") instead (~21 trading days), which comfortably
    covers 1D/2D/3D/1W/2W. Still one Ticker.history() call for all five
    change columns, plus the separate 1d call for the day's actual OHLCV
    bar (kept as literally specified) and the one .info[] pull for the two
    52W fields (also as specified). Deliberate deviation from the literal
    "5d" period string, not an oversight.
  - S_Current_Price / S_Last_Price / S_LastTradedTime: same posture as
    price_action_intraday.py -- sourced from the day's OHLCV bar (Close /
    bar timestamp) rather than a live `.info['currentPrice']` quote, since
    the one permitted `.info[]` pull here is scoped to just the two 52W
    fields per §7b's explicit carve-out.
  - D_Price_vs_52W_High: not present in price_action_4h.py's D_ set:
    added here per the Data Dictionary (§20) formula
    (S_Last_Price - S_52W_High) / S_52W_High, since S_52W_High is now fed
    by this script (and by fetch_engine_weekly.py) and the derived column
    depends on it. True fraction (not pre-scaled), same convention as
    D_Price_vs_Cap_Mid.
  - Volume-anchor state machine (D_Cap_Mid_Anchor / D_Cap_Anchor_Date /
    D_Cap_Anchor_Active / D_Volume_Flag / D_Context_Flag) and its
    threshold/window constants are unchanged from price_action_4h.py --
    logic ported as-is, read from the Legend tab's CONFIG THRESHOLDS block
    at run() start (Sprint 2 Run 5 fix #5 convention).
  - CSV-ingestion trigger: this script does its own save-with-increment
    first, then checks Inputs/HL /  Inputs/BC for *.csv presence. If found,
    portfolio_csv_ingestion.run() is called with no explicit workbook_path
    when this script resolved its own workbook by glob (so the ingestion
    script picks up the version *this* run just saved), or with the same
    explicit path when a test harness passed one in (so tests can inspect
    a single scratch file end-to-end). This means a CSV-triggering EOD run
    produces two incremented workbook versions in one execution (price
    fetch, then ingestion) -- each script's own versioning convention is
    left intact rather than trying to merge the two saves into one.

openpyxl constraint (§1): surgical cell edits only, never a pandas
rewrite of the sheet -- column dtypes/formatting must survive.
"""

import numpy as np
import yfinance as yf
from datetime import datetime
from pathlib import Path

import openpyxl

from workbook_io import (
    find_workbook,
    save_workbook_with_increment,
    write_cell,
    read_legend_scalars,
    INPUTS_HL,
    INPUTS_BC,
    NUMBER_FORMAT_NUMBER,
    NUMBER_FORMAT_PERCENT_SCALED,
    NUMBER_FORMAT_PERCENT_FRACTION,
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEET_NAME = "Universe"

NO_TOUCH = "No Touch"

LEGEND_THRESHOLD_KEYS = [
    "VOL_EXTREME_THRESHOLD",
    "VOL_SPIKE_THRESHOLD",
    "ANCHOR_OVERWRITE_WINDOW_DAYS",
    "ANCHOR_EXPIRE_WINDOW_DAYS",
]

S_COLUMNS_INTRADAY = [
    "S_Daily_High", "S_Daily_Low", "S_Daily_Open",
    "S_Last_Price", "S_Current_Price", "S_LastTradedTime",
    "S_1D_Change_%", "S_2D_Change_%", "S_3D_Change_%",
]

S_COLUMNS_EOD_ONLY = [
    "S_Daily_Close", "S_LastClose_Volume",
    "S_1W_Change_%", "S_2W_Change_%",
    "S_52W_High", "S_52W_Low",
]

S_COLUMNS_ALL = S_COLUMNS_INTRADAY + S_COLUMNS_EOD_ONLY

D_COLUMNS_EOD = [
    "D_Volume_Delta_%", "D_Volume_Flag", "D_Cap_Candle_Mid", "D_Close_Direction",
    "D_Context_Flag", "D_Cap_Mid_Anchor", "D_Cap_Anchor_Date",
    "D_Cap_Anchor_Active", "D_Price_vs_Cap_Mid", "D_Price_vs_52W_High",
]

FIELD_TYPES = {
    "S_Daily_High": "number",
    "S_Daily_Low": "number",
    "S_Daily_Open": "number",
    "S_Last_Price": "number",
    "S_Current_Price": "number",
    "S_LastTradedTime": "text",
    "S_1D_Change_%": "percent_scaled",
    "S_2D_Change_%": "percent_scaled",
    "S_3D_Change_%": "percent_scaled",
    "S_Daily_Close": "number",
    "S_LastClose_Volume": "number",
    "S_1W_Change_%": "percent_scaled",
    "S_2W_Change_%": "percent_scaled",
    "S_52W_High": "number",
    "S_52W_Low": "number",
    "D_Volume_Delta_%": "percent_scaled",
    "D_Volume_Flag": "text",
    "D_Cap_Candle_Mid": "number",
    "D_Close_Direction": "text",
    "D_Context_Flag": "text",
    "D_Cap_Mid_Anchor": "number",
    "D_Cap_Anchor_Active": "text",
    "D_Price_vs_Cap_Mid": "percent_fraction",
    "D_Price_vs_52W_High": "percent_fraction",
}

_FIELD_TYPE_FORMATS = {
    "number": NUMBER_FORMAT_NUMBER,
    "percent_scaled": NUMBER_FORMAT_PERCENT_SCALED,
    "percent_fraction": NUMBER_FORMAT_PERCENT_FRACTION,
    "text": None,
}


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

def header_map(ws):
    """{header_name: column_index} from row 1."""
    return {cell.value: cell.column for cell in ws[1] if cell.value is not None}


def get_cell(ws, headers, row_idx, col_name, default=None):
    col = headers.get(col_name)
    if col is None:
        return default
    val = ws.cell(row=row_idx, column=col).value
    return val if val is not None else default


def set_cell(ws, headers, row_idx, col_name, value):
    col = headers.get(col_name)
    if col is None:
        return  # column not present in sheet -- skip silently, don't crash a run
    field_type = FIELD_TYPES.get(col_name, "text")
    write_cell(ws, row_idx, col, value, number_format=_FIELD_TYPE_FORMATS.get(field_type))


# ---------------------------------------------------------------------------
# Scope filter -- §7b
# ---------------------------------------------------------------------------

def in_scope(ws, headers, row_idx):
    """M_Eliminated != "No Touch" -- all rows, all sleeves, no cadence filter."""
    return get_cell(ws, headers, row_idx, "M_Eliminated") != NO_TOUCH


def resolve_ticker(ws, headers, row_idx):
    yf_ticker = get_cell(ws, headers, row_idx, "S_YF_Ticker")
    if yf_ticker:
        return str(yf_ticker)
    m_ticker = get_cell(ws, headers, row_idx, "M_Ticker")
    return str(m_ticker) if m_ticker else None


# ---------------------------------------------------------------------------
# yfinance fetch -- S_ columns
# ---------------------------------------------------------------------------

def fetch_price_action(symbol):
    """
    One yfinance Ticker object per row: history(1mo) for change-% calcs,
    history(1d) for the day's OHLCV bar, .info[] for the two 52W fields
    only. Returns a dict of the 15 S_ columns fed by §7b, or None if the
    pull failed / no data.
    """
    try:
        t = yf.Ticker(symbol)
        hist_1mo = t.history(period="1mo")
        hist_1d = t.history(period="1d")
        info = t.info
    except Exception as exc:
        print(f"[SKIP] {symbol}: yfinance error: {exc}")
        return None

    if hist_1d.empty or hist_1mo.empty:
        print(f"[SKIP] {symbol}: empty history")
        return None

    last = hist_1d.iloc[-1]
    closes = hist_1mo["Close"]
    last_price = float(last["Close"])

    def pct_change(n):
        if len(closes) <= n:
            return None
        prior = closes.iloc[-1 - n]
        if prior in (0, None) or (isinstance(prior, float) and np.isnan(prior)):
            return None
        return (last_price - prior) / prior * 100

    return {
        "S_Daily_High": float(last["High"]),
        "S_Daily_Low": float(last["Low"]),
        "S_Daily_Open": float(last["Open"]),
        "S_Daily_Close": float(last["Close"]),
        "S_LastClose_Volume": float(last["Volume"]),
        "S_Last_Price": last_price,
        "S_Current_Price": last_price,
        "S_LastTradedTime": hist_1d.index[-1].strftime("%Y-%m-%d %H:%M:%S"),
        "S_1D_Change_%": pct_change(1),
        "S_2D_Change_%": pct_change(2),
        "S_3D_Change_%": pct_change(3),
        "S_1W_Change_%": pct_change(5),
        "S_2W_Change_%": pct_change(10),
        "S_52W_High": info.get("fiftyTwoWeekHigh") if info else None,
        "S_52W_Low": info.get("fiftyTwoWeekLow") if info else None,
    }


# ---------------------------------------------------------------------------
# D_ derived columns (§6 Volume Analysis Layer + §20) -- ported from
# price_action_4h.py, plus D_Price_vs_52W_High
# ---------------------------------------------------------------------------

def business_days_between(start_date, end_date):
    """Simple Mon-Fri trading-day count; holidays not modelled."""
    if start_date is None:
        return None
    if isinstance(start_date, str):
        start_date = datetime.strptime(start_date, "%Y-%m-%d")
    if isinstance(end_date, str):
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
    if end_date < start_date:
        return 0
    days = np.busday_count(start_date.date(), end_date.date())
    return int(days)


def compute_derived(s_values, existing, average_volume, run_date, thresholds):
    """
    s_values: dict from fetch_price_action (this run's S_ columns)
    existing: dict with current D_Cap_Mid_Anchor / D_Cap_Anchor_Date /
              D_Cap_Anchor_Active read from the sheet before this run
    average_volume: S_Average_Volume already on the sheet (from
                    fetch_engine_weekly.py) -- blank/None if not yet
                    populated
    run_date: datetime for "today" (the EOD run date)
    thresholds: dict with VOL_EXTREME_THRESHOLD, VOL_SPIKE_THRESHOLD,
                ANCHOR_OVERWRITE_WINDOW_DAYS, ANCHOR_EXPIRE_WINDOW_DAYS
                (read from the Legend tab by the caller)

    Returns dict of the 10 D_ columns.
    """
    vol_extreme_threshold = thresholds["VOL_EXTREME_THRESHOLD"]
    vol_spike_threshold = thresholds["VOL_SPIKE_THRESHOLD"]
    anchor_overwrite_window_days = thresholds["ANCHOR_OVERWRITE_WINDOW_DAYS"]
    anchor_expire_window_days = thresholds["ANCHOR_EXPIRE_WINDOW_DAYS"]

    high, low = s_values["S_Daily_High"], s_values["S_Daily_Low"]
    open_, close = s_values["S_Daily_Open"], s_values["S_Daily_Close"]
    volume = s_values["S_LastClose_Volume"]

    cap_mid = (high + low) / 2
    close_direction = "UP" if close > open_ else "DOWN"

    if average_volume:
        volume_delta_pct = (volume - average_volume) / average_volume * 100
    else:
        volume_delta_pct = None

    if volume_delta_pct is None:
        volume_flag = None
    elif volume_delta_pct >= vol_extreme_threshold:
        volume_flag = "VOL_EXTREME"
    elif volume_delta_pct >= vol_spike_threshold:
        volume_flag = "VOL_SPIKE"
    else:
        volume_flag = None

    if volume_flag == "VOL_EXTREME" and close_direction == "DOWN":
        context_flag = "CAPITULATION_WATCH"
    elif volume_flag == "VOL_EXTREME" and close_direction == "UP":
        context_flag = "DISTRIBUTION_WATCH"
    else:
        context_flag = None

    price_vs_cap_mid = (close - cap_mid) / cap_mid if cap_mid else None

    # --- Anchor persistence state machine (§6) ---
    anchor = existing.get("D_Cap_Mid_Anchor")
    anchor_date = existing.get("D_Cap_Anchor_Date")
    anchor_active = existing.get("D_Cap_Anchor_Active") == "Y"

    if anchor_active:
        days_since = business_days_between(anchor_date, run_date)
        if days_since is not None and days_since > anchor_expire_window_days:
            anchor, anchor_date, anchor_active = None, None, False
        elif days_since is not None and days_since > anchor_overwrite_window_days:
            anchor, anchor_date, anchor_active = None, None, False
        else:
            if volume_flag == "VOL_EXTREME":
                prior_magnitude = existing.get("_anchor_magnitude")
                new_magnitude = volume_delta_pct
                if prior_magnitude is None or (
                    new_magnitude is not None and new_magnitude > prior_magnitude
                ):
                    anchor = cap_mid
                    anchor_date = run_date.strftime("%Y-%m-%d")
                    anchor_active = True
    else:
        if volume_flag == "VOL_EXTREME":
            anchor = cap_mid
            anchor_date = run_date.strftime("%Y-%m-%d")
            anchor_active = True

    # --- D_Price_vs_52W_High (§20) ---
    high_52w = s_values.get("S_52W_High")
    price_vs_52w_high = (
        (s_values["S_Last_Price"] - high_52w) / high_52w if high_52w else None
    )

    return {
        "D_Volume_Delta_%": volume_delta_pct,
        "D_Volume_Flag": volume_flag,
        "D_Cap_Candle_Mid": cap_mid,
        "D_Close_Direction": close_direction,
        "D_Context_Flag": context_flag,
        "D_Cap_Mid_Anchor": anchor,
        "D_Cap_Anchor_Date": anchor_date,
        "D_Cap_Anchor_Active": "Y" if anchor_active else None,
        "D_Price_vs_Cap_Mid": price_vs_cap_mid,
        "D_Price_vs_52W_High": price_vs_52w_high,
    }


# ---------------------------------------------------------------------------
# CSV ingestion trigger -- §7b
# ---------------------------------------------------------------------------

def csv_inputs_present():
    """True if any *.csv file sits in Inputs/HL or Inputs/BC right now."""
    hl_files = list(INPUTS_HL.glob("*.csv")) if INPUTS_HL.exists() else []
    bc_files = list(INPUTS_BC.glob("*.csv")) if INPUTS_BC.exists() else []
    return bool(hl_files or bc_files)


# ---------------------------------------------------------------------------
# Main run
# ---------------------------------------------------------------------------

def run(workbook_path=None, run_date=None):
    """
    workbook_path: if omitted, the live .xlsm is resolved by globbing the
    base path (aborts on 0 matches) and the save at the end rotates the
    patch-digit filename, moving the old one to Archive/Workbook/. If an
    explicit path is passed (e.g. by the test harness), it's used as-is
    and saved in place.
    """
    run_date = run_date or datetime.now()
    resolved_by_glob = workbook_path is None
    workbook_path = find_workbook() if resolved_by_glob else Path(workbook_path)

    wb = openpyxl.load_workbook(workbook_path, keep_vba=str(workbook_path).endswith(".xlsm"))
    ws = wb[SHEET_NAME]
    headers = header_map(ws)

    legend_scalars = read_legend_scalars(wb, LEGEND_THRESHOLD_KEYS)
    thresholds = {k: float(v) for k, v in legend_scalars.items()}

    processed, skipped = 0, 0
    for row_idx in range(2, ws.max_row + 1):
        if not in_scope(ws, headers, row_idx):
            continue

        symbol = resolve_ticker(ws, headers, row_idx)
        if not symbol:
            skipped += 1
            continue

        s_values = fetch_price_action(symbol)
        if s_values is None:
            skipped += 1
            continue

        for col_name in S_COLUMNS_ALL:
            set_cell(ws, headers, row_idx, col_name, s_values[col_name])

        existing = {
            "D_Cap_Mid_Anchor": get_cell(ws, headers, row_idx, "D_Cap_Mid_Anchor"),
            "D_Cap_Anchor_Date": get_cell(ws, headers, row_idx, "D_Cap_Anchor_Date"),
            "D_Cap_Anchor_Active": get_cell(ws, headers, row_idx, "D_Cap_Anchor_Active"),
            "_anchor_magnitude": get_cell(ws, headers, row_idx, "D_Volume_Delta_%"),
        }
        average_volume = get_cell(ws, headers, row_idx, "S_Average_Volume")
        d_values = compute_derived(s_values, existing, average_volume, run_date, thresholds)
        for col_name in D_COLUMNS_EOD:
            set_cell(ws, headers, row_idx, col_name, d_values[col_name])

        processed += 1

    if resolved_by_glob:
        workbook_path = save_workbook_with_increment(wb, workbook_path)
    else:
        wb.save(workbook_path)

    print(f"Run complete (EOD). {processed} row(s) updated, {skipped} skipped. Workbook: {workbook_path.name}")

    # --- §7b: trigger CSV ingestion if any input files are present ---
    if csv_inputs_present():
        import portfolio_csv_ingestion
        print("  CSV input(s) detected in Inputs/HL or Inputs/BC -- running portfolio_csv_ingestion.run()")
        portfolio_csv_ingestion.run(workbook_path=None if resolved_by_glob else workbook_path)
    else:
        print("  No CSV inputs present in Inputs/HL or Inputs/BC -- skipping ingestion.")

    return wb


if __name__ == "__main__":
    run()
