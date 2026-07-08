"""
One-time Legend-tab migration — Sprint 2 Run 5 fix #5 (+ fix #6).

Adds three new labeled blocks to the Legend tab so portfolio_csv_ingestion.py,
price_action_4h.py and fetch_engine_monthly.py can read their config from the
workbook instead of from hardcoded Python constants:

    CONFIG THRESHOLDS   -- GAP_DUE_DAYS, VOL_EXTREME_THRESHOLD,
                            VOL_SPIKE_THRESHOLD, ANCHOR_OVERWRITE_WINDOW_DAYS,
                            ANCHOR_EXPIRE_WINDOW_DAYS
    COUNTRY LOOKUP      -- raw country name/code -> S_Country abbreviation
                           (includes fix #6: "GB" -> "UK")
    SECTOR ETF LOOKUP   -- Barchart sector text -> S_Sector ETF abbreviation

Idempotent: if a block's header label already exists in column A, that block
is left untouched and skipped (so re-running this against a workbook that
already has the blocks -- e.g. after copying a fresh export from GDrive that
predates this migration -- won't duplicate rows). Run once per workbook that
needs it:

    python setup_legend_config.py

Uses workbook_io.find_workbook() / save_workbook_with_increment() so it
participates in the same version-increment + Archive/Workbook move as the
three pipeline scripts.
"""

from workbook_io import find_workbook, save_workbook_with_increment
import openpyxl

LEGEND_SHEET = "Legend"

CONFIG_THRESHOLDS_LABEL = "CONFIG THRESHOLDS"
COUNTRY_LOOKUP_LABEL = "COUNTRY LOOKUP"
SECTOR_ETF_LOOKUP_LABEL = "SECTOR ETF LOOKUP"

CONFIG_THRESHOLDS_ROWS = [
    ("Key", "Value", None, "Notes"),
    ("GAP_DUE_DAYS", 5, None, "Barchart / Ticker-enrichment gap due-date offset (days) — §10"),
    ("VOL_EXTREME_THRESHOLD", 400, None, "Volume Analysis Layer — §6"),
    ("VOL_SPIKE_THRESHOLD", 200, None, "Volume Analysis Layer — §6"),
    ("ANCHOR_OVERWRITE_WINDOW_DAYS", 5, None, "D_Cap_Mid_Anchor overwrite window (trading days) — §6"),
    ("ANCHOR_EXPIRE_WINDOW_DAYS", 15, None, "D_Cap_Mid_Anchor expire window (trading days) — §6"),
]

# fix #6: "GB" -> "UK" added alongside the existing "United Kingdom" -> "UK"
# mapping -- some sources report the ISO alpha-2 code rather than the full
# country name.
COUNTRY_LOOKUP_ROWS = [
    ("Raw value (yfinance/BC)", "S_Country abbreviation"),
    ("United States", "US"),
    ("United Kingdom", "UK"),
    ("GB", "UK"),
    ("Germany", "DE"),
    ("France", "FR"),
    ("Ireland", "IE"),
    ("Netherlands", "NL"),
    ("Switzerland", "CH"),
    ("Denmark", "DK"),
    ("Sweden", "SE"),
    ("Spain", "ES"),
    ("Italy", "IT"),
    ("Japan", "JP"),
    ("Canada", "CA"),
    ("Australia", "AU"),
    ("Hong Kong", "HK"),
    ("Singapore", "SG"),
]

SECTOR_ETF_LOOKUP_ROWS = [
    ("Barchart Sector text", "S_Sector ETF"),
    ("Technology", "XLK"),
    ("Oils-Energy", "XLE"),
    ("Consumer Discretionary", "XLY"),
    ("Consumer Staples", "XLP"),
    ("Healthcare", "XLV"),
    ("Financials", "XLF"),
    ("Finance", "XLF"),
    ("Industrials", "XLI"),
    ("Materials", "XLB"),
    ("Real Estate", "XLRE"),
    ("Utilities", "XLU"),
    ("Communication Services", "XLC"),
    ("Energy", "XLE"),
]


def _label_exists(ws, label):
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == label:
            return True
    return False


def _append_block(ws, label, data_rows):
    if _label_exists(ws, label):
        print(f"  [SKIP] '{label}' block already present.")
        return
    start_row = ws.max_row + 2  # blank separator line
    ws.cell(row=start_row, column=1, value=label)
    for offset, row_values in enumerate(data_rows, start=1):
        for col_offset, value in enumerate(row_values, start=1):
            if value is not None:
                ws.cell(row=start_row + offset, column=col_offset, value=value)
    print(f"  [OK] '{label}' block added at row {start_row}.")


def run(workbook_path=None):
    resolved_by_glob = workbook_path is None
    workbook_path = find_workbook() if resolved_by_glob else workbook_path

    wb = openpyxl.load_workbook(workbook_path, keep_vba=True)
    ws = wb[LEGEND_SHEET]

    print(f"Migrating Legend tab in {workbook_path}")
    _append_block(ws, CONFIG_THRESHOLDS_LABEL, CONFIG_THRESHOLDS_ROWS)
    _append_block(ws, COUNTRY_LOOKUP_LABEL, COUNTRY_LOOKUP_ROWS)
    _append_block(ws, SECTOR_ETF_LOOKUP_LABEL, SECTOR_ETF_LOOKUP_ROWS)

    if resolved_by_glob:
        workbook_path = save_workbook_with_increment(wb, workbook_path)
    else:
        wb.save(workbook_path)
    print(f"Saved: {workbook_path}")
    return workbook_path


if __name__ == "__main__":
    run()
