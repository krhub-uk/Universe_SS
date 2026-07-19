"""
Fetch Engine -- Weekly cadence (fundamentals)
Spec: 00_Portfolio_Automation_Spec_V5.5.md, Section 7c. Sprint 4.

Runs: Sunday (weekly), after price_action_eod.py completes (run_weekly.sh).

Scope (checked in this order):
    M_Eliminated != "No Touch"          (universal kill switch, checked first)
    M_Div_Coupon_Class in ELIGIBLE_DIV_COUPON_CLASSES
        {Aristocrat_King, Aristocrat, Achiever, Contender, HighIncome, MedIncome}

S_ columns fed -- direct 1:1 off a single yfinance .info pull:
    S_Name, S_Country, S_Exchange, S_Industry, S_Sub_Industry,
    S_MCap, S_Beta, S_PE_Ratio, S_PayoutRatio_Pct, S_DebtEquity, S_ROE,
    S_Dividend_Yield_Pct, S_Average_Volume,
    S_ExDividend_Date, S_Reporting_Date, S_LastTradedTime

S_ columns removed from weekly (moved to EOD cadence):
    S_52W_High, S_52W_Low  -- see price_action_eod.py

S_Sector authority: populate_sector_from_lookups() runs before all fetch loops.
Reads Lookups tab M_Ticker/S_Sector columns. No yfinance sector fetch.
S_BC_Sector remains -- written by Barchart ingestion, untouched here.

S_ columns fed -- computed 5Y CAGR / averages:
    S_EPS_Growth_5Y_Pct, S_DivGrowth_5Y_Pct, S_DivYield_5YAvg_Pct, S_PE_5YAvg,
    S_Yrs_DivIncome_Buys_1Share

S_ columns fed -- Wave 6:
    S_Price_5Y_Return_Pct: yfinance history(period="5y", auto_adjust=True),
    (last_close - first_close) / first_close * 100. Blank if < 200 rows.

Percentage storage (Sprint 4 / _Pct rename):
    S_PayoutRatio_Pct: yfinance returns fraction (0.64) -> stored 64.0
    S_Dividend_Yield_Pct: yfinance returns fraction (0.04) -> stored 4.0
    S_DivGrowth_5Y_Pct, S_DivYield_5YAvg_Pct, S_EPS_Growth_5Y_Pct: already x100
    All _Pct columns use General cell format (no Excel % format).

Gap logging: Outputs/BC/fetch_gaps_YYYYMMDD.csv (unchanged).
Wave 8: Structured PHASE/ACTION/TICKER/RUN_ID/UID logging.
"""

import csv
import logging
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import yfinance as yf
import openpyxl

from workbook_io import (
    BASE_PATH,
    find_workbook,
    save_workbook_with_increment,
    OUTPUTS_BC,
    write_cell,
    read_legend_lookup_table,
    NUMBER_FORMAT_NUMBER,
    NUMBER_FORMAT_PERCENT_SCALED,
    NUMBER_FORMAT_PERCENT_FRACTION,
    NUMBER_FORMAT_DATE,
)
from icu_client import push_status, check_gate, resolve_version

COMPONENT_ID = "universe_ss_weekly"

# ---------------------------------------------------------------------------
# Logging (Wave 8)
# ---------------------------------------------------------------------------

LOG_DIR = BASE_PATH / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "fetch_engine_weekly.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("fetch_engine_weekly")

_ctx: dict = {}
_VOCAB_KEYS = ("PHASE=", "ACTION=", "TICKER=", "RUN_ID=", "UID=")


def _log(level: str, phase: str, action: str, ticker: str, message: str) -> None:
    run_id = _ctx.get("run_id", "")
    uid    = _ctx.get("uid", "")
    line   = (
        f"PHASE={phase} ACTION={action} TICKER={ticker} "
        f"RUN_ID={run_id} UID={uid} | {message}"
    )
    for key in _VOCAB_KEYS:
        if key not in line:
            log.warning(f"[VOCAB_FAIL] Missing {key}")
    getattr(log, level.lower())(line)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

SHEET_NAME = "Universe"

NO_TOUCH = "No Touch"

ELIGIBLE_DIV_COUPON_CLASSES = {
    "Aristocrat_King", "Aristocrat", "Achiever", "Contender", "HighIncome", "MedIncome",
}

GAP_OUTPUT_DIR = OUTPUTS_BC

# Wave 4: S_52W_High/Low removed (moved to EOD). S_Sector removed (from Lookups).
# Wave 4: _Pct renames. Wave 6: S_Price_5Y_Return_Pct added.
S_COLUMNS_INFO_DIRECT = [
    "S_Name", "S_Country", "S_Exchange", "S_Industry", "S_Sub_Industry",
    "S_MCap", "S_Beta", "S_PE_Ratio", "S_PayoutRatio_Pct", "S_DebtEquity", "S_ROE",
    "S_Dividend_Yield_Pct", "S_Average_Volume",
    "S_ExDividend_Date", "S_Reporting_Date", "S_LastTradedTime",
]

S_COLUMNS_COMPUTED = [
    "S_EPS_Growth_5Y_Pct", "S_DivGrowth_5Y_Pct", "S_DivYield_5YAvg_Pct", "S_PE_5YAvg",
    "S_Yrs_DivIncome_Buys_1Share",
    "S_Price_5Y_Return_Pct",  # Wave 6
]

S_COLUMNS_ALL = S_COLUMNS_INFO_DIRECT + S_COLUMNS_COMPUTED

# All _Pct columns use "number" type (General cell format per spec).
# percent_fraction kept only for S_ROE (not renamed, still fraction from yfinance).
FIELD_TYPES = {
    "S_Name":                  "text",
    "S_Country":               "text",
    "S_Exchange":              "text",
    "S_Industry":              "text",
    "S_Sub_Industry":          "text",
    "S_MCap":                  "number",
    "S_Beta":                  "number",
    "S_PE_Ratio":              "number",
    "S_PayoutRatio_Pct":       "number",   # stored x100, General format
    "S_DebtEquity":            "number",
    "S_ROE":                   "percent_fraction",
    "S_Dividend_Yield_Pct":    "number",   # stored x100, General format
    "S_Average_Volume":        "number",
    "S_ExDividend_Date":       "date",
    "S_Reporting_Date":        "date",
    "S_LastTradedTime":        "text",
    "S_EPS_Growth_5Y_Pct":     "number",   # stored x100
    "S_DivGrowth_5Y_Pct":      "number",   # stored x100
    "S_DivYield_5YAvg_Pct":    "number",   # stored x100
    "S_PE_5YAvg":              "number",
    "S_Yrs_DivIncome_Buys_1Share": "number",
    "S_Price_5Y_Return_Pct":   "number",   # stored x100
}

_FIELD_TYPE_FORMATS = {
    "number":           NUMBER_FORMAT_NUMBER,
    "percent_scaled":   NUMBER_FORMAT_PERCENT_SCALED,
    "percent_fraction": NUMBER_FORMAT_PERCENT_FRACTION,
    "date":             NUMBER_FORMAT_DATE,
    "text":             None,
}


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------

def header_map(ws):
    """{header_name: column_index} from row 1."""
    return {cell.value: cell.column for cell in ws[1] if cell.value is not None}


def get_cell(ws, headers, row_idx, col_name, default=None):
    col = headers.get(col_name)
    if col is None:
        return default
    val = ws.cell(row=row_idx, column=col).value
    return val if val is not None else default


def set_cell(ws, headers, row_idx, col_name, value):
    col = headers.get(col_name)
    if col is None:
        return  # column not in sheet -- skip silently
    field_type = FIELD_TYPES.get(col_name, "text")
    write_cell(ws, row_idx, col, value, number_format=_FIELD_TYPE_FORMATS.get(field_type))


# ---------------------------------------------------------------------------
# S_Sector from Lookups (Wave 4)
# ---------------------------------------------------------------------------

def populate_sector_from_lookups(wb, ws_universe):
    """
    Sprint 4 Wave 4: S_Sector authority is Lookups tab, not yfinance.
    Reads Lookups tab columns M_Ticker (col A) and S_Sector (col E) by name.
    For each Universe row: looks up M_Ticker in Lookups and writes S_Sector.
    No match = blank. No error raised. Runs before all fetch loops.
    """
    ws_lookups = wb["Lookups"]
    lk_headers = header_map(ws_lookups)

    ticker_col = lk_headers.get("M_Ticker")
    sector_col = lk_headers.get("S_Sector")
    if ticker_col is None or sector_col is None:
        _log("warning", "SECTOR", "LOOKUPS_COLS_MISSING", "",
             "Lookups tab missing M_Ticker or S_Sector column -- sector populate skipped")
        return

    # Build ticker -> sector dict from Lookups
    sector_map = {}
    for row in ws_lookups.iter_rows(min_row=2, values_only=True):
        tk = row[ticker_col - 1]
        sc = row[sector_col - 1]
        if tk:
            sector_map[str(tk).strip()] = str(sc).strip() if sc else ""

    uni_headers = header_map(ws_universe)
    updated = 0
    for row_idx in range(2, ws_universe.max_row + 1):
        m_ticker = get_cell(ws_universe, uni_headers, row_idx, "M_Ticker")
        if not m_ticker:
            continue
            sector = sector_map.get(str(m_ticker).strip())
            if sector:
                set_cell(ws_universe, uni_headers, row_idx, "S_Sector", sector)
                #sector = sector_map.get(str(m_ticker).strip(), "")
                #set_cell(ws_universe, uni_headers, row_idx, "S_Sector", sector)
        
        if sector:
            updated += 1

    _log("info", "SECTOR", "LOOKUPS_POPULATE", "",
         f"S_Sector populated from Lookups for {updated} rows")


# ---------------------------------------------------------------------------
# Scope filter
# ---------------------------------------------------------------------------

def in_scope(ws, headers, row_idx):
    if get_cell(ws, headers, row_idx, "M_Eliminated") == NO_TOUCH:
        return False
    div_class = get_cell(ws, headers, row_idx, "M_Div_Coupon_Class")
    return div_class in ELIGIBLE_DIV_COUPON_CLASSES


def resolve_ticker(ws, headers, row_idx):
    yf_ticker = get_cell(ws, headers, row_idx, "S_YF_Ticker")
    if yf_ticker:
        return str(yf_ticker)
    m_ticker = get_cell(ws, headers, row_idx, "M_Ticker")
    return str(m_ticker) if m_ticker else None


# ---------------------------------------------------------------------------
# yfinance fetch -- direct info[] columns
# ---------------------------------------------------------------------------

def _unix_to_date_str(ts):
    if ts is None:
        return None
    try:
        from datetime import timezone
        return datetime.fromtimestamp(ts, tz=timezone.utc).date()
    except (TypeError, ValueError, OSError):
        return None


def _unix_to_datetime_str(ts):
    if ts is None:
        return None
    try:
        from datetime import timezone
        return datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError, OSError):
        return None


def _fraction_to_pct(v):
    """Convert yfinance fraction (0.04) to x100 storage (4.0). None -> None."""
    if v is None:
        return None
    try:
        return float(v) * 100
    except (TypeError, ValueError):
        return None


def fetch_info_fields(info, country_lookup):
    """
    Direct 1:1 mapping off a single yfinance .info dict pull.
    Wave 4: S_Sector removed (from Lookups now). S_52W_High/Low removed (EOD).
    Wave 4: S_PayoutRatio_Pct and S_Dividend_Yield_Pct stored x100.
    """
    from workbook_io import read_legend_lookup_table  # keep import local to avoid circular

    def map_country(raw_country):
        if raw_country is None:
            return None
        return country_lookup.get(raw_country, raw_country)

    return {
        "S_Name":               info.get("longName") or info.get("shortName"),
        "S_Country":            map_country(info.get("country")),
        "S_Exchange":           info.get("exchange"),
        "S_Industry":           info.get("industry"),
        "S_Sub_Industry":       info.get("industryDisp") or info.get("industry"),
        "S_MCap":               info.get("marketCap"),
        "S_Beta":               info.get("beta"),
        "S_PE_Ratio":           info.get("trailingPE"),
        # Wave 4: stored x100 (yfinance returns fraction 0.64 -> store 64.0)
        "S_PayoutRatio_Pct":    _fraction_to_pct(info.get("payoutRatio")),
        "S_DebtEquity":         info.get("debtToEquity"),
        "S_ROE":                info.get("returnOnEquity"),
        # Wave 4: stored x100 (yfinance returns fraction 0.04 -> store 4.0)
        "S_Dividend_Yield_Pct": _fraction_to_pct(info.get("dividendYield")),
        "S_Average_Volume":     info.get("averageVolume"),
        "S_ExDividend_Date":    _unix_to_date_str(info.get("exDividendDate")),
        "S_Reporting_Date":     _unix_to_date_str(
            info.get("earningsTimestamp") or info.get("mostRecentQuarter")
        ),
        "S_LastTradedTime":     _unix_to_datetime_str(info.get("regularMarketTime")),
    }


# ---------------------------------------------------------------------------
# yfinance fetch -- computed 5Y CAGR / average columns
# ---------------------------------------------------------------------------

def cagr(start_value, end_value, years):
    if not start_value or not end_value or start_value <= 0 or end_value <= 0 or not years:
        return None
    try:
        return ((end_value / start_value) ** (1 / years) - 1) * 100
    except (ZeroDivisionError, ValueError):
        return None


def fetch_computed_fields(ticker_obj, info):
    """
    Best-effort 5Y CAGR / average fields + S_Price_5Y_Return_Pct (Wave 6).
    Wave 4: all _Growth/_Yield computed columns renamed to _Pct variants.
    Returns None for any field where insufficient history exists.
    """
    result = {col: None for col in S_COLUMNS_COMPUTED}
    current_year = datetime.now().year

    # --- EPS growth 5Y (stored x100 via cagr()) ---
    try:
        income = ticker_obj.income_stmt
        if income is not None and not income.empty:
            eps_row = None
            for label in ("Diluted EPS", "Basic EPS"):
                if label in income.index:
                    eps_row = income.loc[label].dropna()
                    break
            if eps_row is not None and len(eps_row) >= 2:
                eps_row = eps_row.sort_index()
                span_years = len(eps_row) - 1
                result["S_EPS_Growth_5Y_Pct"] = cagr(eps_row.iloc[0], eps_row.iloc[-1], span_years)
    except Exception:
        pass

    # --- Dividend growth 5Y + Dividend yield 5Y avg ---
    annual_divs = None
    try:
        dividends = ticker_obj.dividends
        if dividends is not None and not dividends.empty:
            annual = dividends.groupby(dividends.index.year).sum()
            annual = annual[annual.index < current_year]
            annual_divs = annual.tail(6)
            if len(annual_divs) >= 2:
                span_years = annual_divs.index[-1] - annual_divs.index[0]
                result["S_DivGrowth_5Y_Pct"] = cagr(
                    annual_divs.iloc[0], annual_divs.iloc[-1], span_years
                )
    except Exception:
        pass

    try:
        if annual_divs is not None and len(annual_divs) >= 1:
            hist = ticker_obj.history(period="6y")
            if hist is not None and not hist.empty:
                hist_annual_price = hist["Close"].groupby(hist.index.year).mean()
                yields = []
                for yr, div_sum in annual_divs.items():
                    price = hist_annual_price.get(yr)
                    if price:
                        yields.append(div_sum / price * 100)
                if yields:
                    result["S_DivYield_5YAvg_Pct"] = sum(yields) / len(yields)
    except Exception:
        pass

    # --- PE 5Y avg ---
    try:
        trailing_eps = info.get("trailingEps")
        hist5 = ticker_obj.history(period="5y")
        if trailing_eps and hist5 is not None and not hist5.empty:
            avg_close = hist5["Close"].mean()
            result["S_PE_5YAvg"] = avg_close / trailing_eps
    except Exception:
        pass

    # --- Yrs of dividend income to buy back 1 share ---
    try:
        price = info.get("currentPrice") or info.get("regularMarketPrice")
        annual_dividend = info.get("dividendRate")
        if price and annual_dividend:
            result["S_Yrs_DivIncome_Buys_1Share"] = price / annual_dividend
    except Exception:
        pass

    # --- S_Price_5Y_Return_Pct (Wave 6) ---
    try:
        hist5y = ticker_obj.history(period="5y", auto_adjust=True)
        if hist5y is not None and not hist5y.empty and len(hist5y) >= 200:
            first_close = hist5y["Close"].iloc[0]
            last_close  = hist5y["Close"].iloc[-1]
            if first_close and first_close != 0:
                result["S_Price_5Y_Return_Pct"] = (last_close - first_close) / first_close * 100
        # else: < 200 rows -> stays None (blank)
    except Exception:
        pass

    return result


# ---------------------------------------------------------------------------
# Combined per-ticker fetch
# ---------------------------------------------------------------------------

def fetch_fundamentals(symbol, country_lookup):
    """
    One yfinance Ticker object per row; .info + .income_stmt + .dividends +
    .history() all pulled from it. Returns (values_dict, gap_rows).
    """
    try:
        t = yf.Ticker(symbol)
        info = t.info
    except Exception as exc:
        _log("warning", "FETCH", "YF_ERROR", symbol, f"yfinance error: {exc}")
        return None, [{"column": "ALL", "reason": f"yfinance error: {exc}"}]

    if not info or (info.get("regularMarketPrice") is None and info.get("currentPrice") is None):
        _log("warning", "FETCH", "YF_EMPTY", symbol, "empty/invalid info payload")
        return None, [{"column": "ALL", "reason": "empty or invalid info payload"}]

    values = fetch_info_fields(info, country_lookup)
    values.update(fetch_computed_fields(t, info))

    gaps = [
        {"column": col, "reason": "not available from yfinance for this ticker"}
        for col, val in values.items()
        if val is None
    ]
    return values, gaps


# ---------------------------------------------------------------------------
# Gap logging
# ---------------------------------------------------------------------------

def log_gaps(gap_rows, run_date, output_dir=GAP_OUTPUT_DIR):
    if not gap_rows:
        return None
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path   = output_dir / f"fetch_gaps_{run_date.strftime('%Y%m%d')}.csv"
    file_exists = out_path.exists()
    fieldnames  = ["ticker", "column", "reason", "detected_date"]
    with open(out_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerows(gap_rows)
    return out_path


# ---------------------------------------------------------------------------
# Main run
# ---------------------------------------------------------------------------

def run(workbook_path=None, run_date=None, gap_output_dir=GAP_OUTPUT_DIR):
    _ctx["run_id"] = str(uuid.uuid4())[:8]
    _ctx["uid"]    = ""

    start = datetime.now()
    run_date = run_date or start
    _log("info", "STARTUP", "RUN_START", "", "fetch_engine_weekly.py starting")

    resolved_by_glob = workbook_path is None
    workbook_path    = find_workbook() if resolved_by_glob else Path(workbook_path)

    wb = openpyxl.load_workbook(workbook_path, keep_vba=str(workbook_path).endswith(".xlsm"))
    ws = wb[SHEET_NAME]
    headers = header_map(ws)

    country_lookup = read_legend_lookup_table(wb, "COUNTRY LOOKUP")

    # Wave 4: populate S_Sector from Lookups BEFORE all fetch loops
    _log("info", "SECTOR", "POPULATE_START", "", "Populating S_Sector from Lookups")
    populate_sector_from_lookups(wb, ws)

    processed, skipped = 0, 0
    all_gaps = []

    for row_idx in range(2, ws.max_row + 1):
        if not in_scope(ws, headers, row_idx):
            continue

        symbol = resolve_ticker(ws, headers, row_idx)
        if not symbol:
            skipped += 1
            continue

        _log("info", "FETCH", "TICKER_START", symbol, f"row {row_idx}")
        values, gaps = fetch_fundamentals(symbol, country_lookup)
        for g in gaps:
            g["ticker"]        = symbol
            g["detected_date"] = run_date.strftime("%Y-%m-%d")
        all_gaps.extend(gaps)

        if values is None:
            _log("warning", "FETCH", "TICKER_SKIP", symbol, "fetch returned None -- skipped")
            skipped += 1
            continue

        for col_name in S_COLUMNS_ALL:
            set_cell(ws, headers, row_idx, col_name, values[col_name])

        _log("info", "FETCH", "TICKER_DONE", symbol, f"row {row_idx} written")
        processed += 1

    gap_path = log_gaps(all_gaps, run_date, gap_output_dir)

    _ctx["processed"] = processed
    _ctx["skipped"] = skipped

    if resolved_by_glob:
        workbook_path = save_workbook_with_increment(wb, workbook_path)
    else:
        wb.save(workbook_path)

    elapsed = str(datetime.now() - start).split(".")[0]
    _log("info", "COMPLETE", "RUN_END", "",
         f"{processed} row(s) updated, {skipped} skipped, {len(all_gaps)} gap(s)"
         + (f", gaps -> {gap_path}" if gap_path else "")
         + f". Duration {elapsed}")

    return wb


if __name__ == "__main__":
    workbook_path = find_workbook()
    VERSION = resolve_version(str(workbook_path))

    if not check_gate(COMPONENT_ID):
        push_status(
            COMPONENT_ID, "PAUSED", VERSION,
            last_run_result="SKIPPED",
            trigger="GATE_CHECK"
        )
        sys.exit(0)

    run_start = datetime.now(timezone.utc).isoformat()
    push_status(
        COMPONENT_ID, "RUNNING", VERSION,
        last_run_utc=run_start,
        trigger="SCHEDULED"
    )

    try:
        run()
        processed = _ctx.get("processed", 0)
        skipped = _ctx.get("skipped", 0)
        if skipped == 0:
            result = "SUCCESS"
        elif processed > 0:
            result = "PARTIAL"
        else:
            result = "DEGRADED"
        push_status(
            COMPONENT_ID, "IDLE", VERSION,
            last_run_utc=run_start,
            last_run_result=result,
            trigger="SCHEDULED",
            metrics={"tickers_fetched": processed, "tickers_failed": skipped}
        )
    except Exception as e:
        push_status(
            COMPONENT_ID, "ERROR", VERSION,
            last_run_utc=run_start,
            last_run_result="FAILED",
            message=str(e)
        )
        raise
